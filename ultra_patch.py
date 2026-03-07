# =========================================================
# FUTURE ULTRA MASTER PATCH
# scalony patch (UI + EMAIL + EXCEL + PREMIUM OVERRIDE)
# kompatybilny z main.py
# =========================================================

import threading
from pathlib import Path
from datetime import datetime

from kivy.metrics import dp
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.animation import Animation
from kivy.utils import get_color_from_hex

from kivy.graphics import Color, RoundedRectangle

from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.widget import Widget
from kivy.uix.filechooser import FileChooserListView


# =========================================================
# GLOBALNE ZMIENNE
# =========================================================

ULTRA_ENABLED = True

UI_PRIMARY = (0.12, 0.45, 0.95, 1)
UI_BG = (0.96, 0.97, 0.99, 1)


# =========================================================
# ULTRA BUTTON
# =========================================================

class UltraButton(Button):

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        self.background_normal = ""
        self.background_color = (0,0,0,0)

        self.color = (1,1,1,1)

        self.size_hint_y = None
        self.height = dp(52)

        self.font_size = 16
        self.bold = True

        with self.canvas.before:

            Color(*UI_PRIMARY)

            self.bg = RoundedRectangle(
                radius=[18],
                pos=self.pos,
                size=self.size
            )

        self.bind(pos=self.update_bg,size=self.update_bg)


    def update_bg(self,*args):

        self.bg.pos=self.pos
        self.bg.size=self.size


    def on_press(self):

        Animation.cancel_all(self)

        anim = Animation(opacity=.85,duration=.08)
        anim += Animation(opacity=1,duration=.08)

        anim.start(self)



# =========================================================
# ULTRA INPUT
# =========================================================

class UltraInput(TextInput):

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        self.font_size = 18
        self.padding = [dp(14), dp(14)]

        self.size_hint_y = None
        self.height = dp(56)

        self.background_normal = ""
        self.background_active = ""

        self.background_color = (1,1,1,1)

        with self.canvas.before:

            Color(.85,.88,.92,1)

            self.bg = RoundedRectangle(
                radius=[12],
                pos=self.pos,
                size=self.size
            )

        self.bind(pos=self.update_bg,size=self.update_bg)


    def update_bg(self,*args):

        self.bg.pos=self.pos
        self.bg.size=self.size



# =========================================================
# ULTRA CARD
# =========================================================

class UltraCard(BoxLayout):

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        self.padding = dp(16)
        self.spacing = dp(10)

        with self.canvas.before:

            Color(1,1,1,1)

            self.bg = RoundedRectangle(
                radius=[20],
                pos=self.pos,
                size=self.size
            )

        self.bind(pos=self.update_bg,size=self.update_bg)


    def update_bg(self,*args):

        self.bg.pos=self.pos
        self.bg.size=self.size



# =========================================================
# ULTRA LAYOUT
# =========================================================

class UltraLayout(BoxLayout):

    def __init__(self, **kwargs):

        super().__init__(**kwargs)

        with self.canvas.before:

            Color(*UI_BG)

            self.bg = RoundedRectangle(
                pos=self.pos,
                size=self.size
            )

        self.bind(pos=self.update_bg,size=self.update_bg)


    def update_bg(self,*args):

        self.bg.pos=self.pos
        self.bg.size=self.size



# =========================================================
# UI LOADER
# =========================================================

class UILoader(Widget):
    pass



# =========================================================
# EMAIL EXCEL IMPORT
# =========================================================

def pro_load_email_excel(self,path):

    from openpyxl import load_workbook

    wb = load_workbook(path,data_only=True)

    sheet = wb.active

    rows=list(sheet.iter_rows(values_only=True))

    header=[str(x).lower() for x in rows[0]]

    try:

        name_i=header.index("imię")
        surname_i=header.index("nazwisko")
        email_i=header.index("email")

    except:

        self.popup(
            "Błąd",
            "Excel musi mieć kolumny:\nImię | Nazwisko | Email"
        )
        return

    self.email_map={}

    for row in rows[1:]:

        name=str(row[name_i]).strip()
        surname=str(row[surname_i]).strip()
        email=str(row[email_i]).strip()

        key=f"{name} {surname}".lower()

        self.email_map[key]=email

    wb.close()

    self.popup(
        "Email",
        f"Wczytano {len(self.email_map)} adresów"
    )



# =========================================================
# EMAIL FILE PICKER
# =========================================================

def pro_email_file_popup(self):

    chooser=FileChooserListView(filters=["*.xlsx"])

    layout=UltraLayout(orientation="vertical")

    btn=UltraButton(text="Wczytaj plik")

    popup=Popup(
        title="Wybierz plik z emailami",
        content=layout,
        size_hint=(0.9,0.9)
    )

    def load(_):

        if not chooser.selection:
            return

        self.load_email_excel(chooser.selection[0])

        popup.dismiss()

    btn.bind(on_press=load)

    layout.add_widget(chooser)
    layout.add_widget(btn)

    popup.open()



# =========================================================
# EMAIL SEARCH
# =========================================================

def pro_find_email(self,row):

    if not hasattr(self,"email_map"):
        return None

    if len(row)<2:
        return None

    name=str(row[0]).strip()
    surname=str(row[1]).strip()

    key=f"{name} {surname}".lower()

    return self.email_map.get(key)



# =========================================================
# SMTP TEST
# =========================================================

def pro_test_smtp(self,_):

    import smtplib

    def run():

        try:

            server=smtplib.SMTP(
                self.smtp_server.text,
                int(self.smtp_port.text),
                timeout=20
            )

            server.starttls()

            server.login(
                self.smtp_user.text,
                self.smtp_pass.text
            )

            server.quit()

            Clock.schedule_once(
                lambda dt:self.popup("SMTP","Połączenie OK")
            )

        except Exception as e:

            Clock.schedule_once(
                lambda dt:self.popup("SMTP ERROR",str(e))
            )

    threading.Thread(target=run).start()



# =========================================================
# EMAIL THREAD
# =========================================================

def pro_email_thread(self):

    import smtplib
    from email.message import EmailMessage

    smtp=self.load_smtp()

    if not smtp:
        return

    try:

        server=smtplib.SMTP(
            smtp["server"],
            int(smtp["port"])
        )

        server.starttls()

        server.login(
            smtp["user"],
            smtp["pass"]
        )

    except Exception as e:

        Clock.schedule_once(
            lambda dt:self.popup("SMTP",str(e))
        )

        return

    rows=self.full_data[1:]

    sent=0

    for row in rows:

        email=pro_find_email(self,row)

        if not email:
            continue

        msg=EmailMessage()

        msg["Subject"]="Informacja"
        msg["From"]=smtp["user"]
        msg["To"]=email

        msg.set_content("Automatyczna wiadomość")

        try:

            server.send_message(msg)
            sent+=1
        except:
            pass

    server.quit()

    Clock.schedule_once(
        lambda dt:self.popup(
            "Email",
            f"Wysłano {sent} wiadomości"
        )
    )



# =========================================================
# ULTRA EXCEL EXPORT
# =========================================================

def ultra_export_excel(self):

    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side

    rows=self.filtered_data

    if not rows:
        return

    folder=Path("/storage/emulated/0/Documents/FutureExport")
    folder.mkdir(parents=True,exist_ok=True)

    bold=Font(bold=True)

    thick=Side(style="thick")

    border=Border(
        left=thick,
        right=thick,
        top=thick,
        bottom=thick
    )

    header=rows[0]

    for row in rows[1:]:

        wb=Workbook()
        ws=wb.active

        ws.append(header)
        ws.append(row)

        for cell in ws[1]:
            cell.font=bold
            cell.border=border

        for r in ws.iter_rows(min_row=2):
            for cell in r:
                cell.border=border

        name=row[0] if row else "file"

        now=datetime.now().strftime("%Y%m%d_%H%M%S")

        file=folder/f"{name}_{now}.xlsx"

        wb.save(file)

    Clock.schedule_once(
        lambda dt:self.popup("Export","Zakończony")
    )



# =========================================================
# GLOBAL UI PATCH
# =========================================================

def apply_ui_patch():

    Window.clearcolor=(0.04,0.05,0.09,1)



# =========================================================
# PREMIUM OVERRIDE
# =========================================================

def override_premium(app_module):

    if hasattr(app_module,"PremiumButton"):

        app_module.PremiumButton=UltraButton



# =========================================================
# APP PATCH
# =========================================================

def patch_app(app):

    if hasattr(app,"title"):
        app.title="Future 9.0 ULTRA"

    app.ultra_mode=True



# =========================================================
# METHOD OVERRIDE
# =========================================================

def override_methods(app):

    if hasattr(app,"popup"):

        original_popup=app.popup

        def ultra_popup(title,text):

            text=f"⚡ ULTRA\n\n{text}"

            return original_popup(title,text)

        app.popup=ultra_popup



# =========================================================
# PODPIĘCIE FUNKCJI DO APP
# =========================================================

def attach_features(app):

    app.load_email_excel = pro_load_email_excel
    app.load_email_list_popup = pro_email_file_popup
    app.find_email_for_row = pro_find_email
    app.test_smtp = pro_test_smtp
    app._email_thread = pro_email_thread
    app.export_excel = ultra_export_excel



# =========================================================
# FULL PATCH LOADER
# =========================================================

def apply_patch(app_module, app_instance):

    apply_ui_patch()

    override_premium(app_module)

    patch_app(app_instance)

    override_methods(app_instance)

    attach_features(app_instance)

    print("⚡ FUTURE ULTRA PATCH LOADED")
