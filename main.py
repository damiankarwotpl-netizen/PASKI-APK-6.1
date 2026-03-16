import os
import json
import sqlite3
import threading
import smtplib
import mimetypes
import time
import random
import traceback
import sys
import webbrowser
import urllib.parse
from datetime import datetime
from pathlib import Path
from email.message import EmailMessage

from kivy.app import App
from kivy.metrics import dp
from kivy.clock import Clock
from kivy.utils import platform
from kivy.core.window import Window
from kivy.core.text import Label as CoreLabel
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.uix.checkbox import CheckBox
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.progressbar import ProgressBar
from kivy.animation import Animation
from kivy.graphics import Color, Rectangle, RoundedRectangle, Line

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
except ImportError:
    load_workbook = Workbook = None
try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from reportlab.pdfgen import canvas
except:
    canvas = None

try:
    import pandas as pd
except Exception:
    pd = None

from collections import defaultdict

try:
    from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    SimpleDocTemplate = None
    Table = None
    Paragraph = None
    Spacer = None
    A4 = None
    getSampleStyleSheet = None

COLOR_PRIMARY = (0.1, 0.5, 0.9, 1)
COLOR_BG = (0.05, 0.07, 0.1, 1)
COLOR_CARD = (0.12, 0.15, 0.2, 1)
COLOR_TEXT = (0.95, 0.95, 0.95, 1)
COLOR_ROW_A = (0.08, 0.1, 0.15, 1)
COLOR_ROW_B = (0.13, 0.16, 0.22, 1)
COLOR_HEADER = (0.1, 0.2, 0.35, 1)

class ModernButton(Button):
    def __init__(self, bg_color=COLOR_PRIMARY, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = ""
        self.background_color = (0,0,0,0)
        self.color = COLOR_TEXT
        self.bold = True
        self.radius = [dp(12)]
        self.base_color = bg_color
        self._font_max = float(getattr(self, 'font_size', dp(16)))
        self._font_min = float(dp(10))
        with self.canvas.before:
            self.bg = Color(*self.base_color)
            self.rect = RoundedRectangle(pos=self.pos, size=self.size, radius=self.radius)
            Color(1, 1, 1, 0.12)
            self.border_line = Line(rounded_rectangle=(self.x, self.y, self.width, self.height, dp(12)), width=1)
        state_handler = getattr(self, '_update_state', None)
        if not callable(state_handler):
            state_handler = self._fallback_update_state
        self.halign = 'center'
        self.valign = 'middle'
        self.shorten = False
        self.max_lines = 1
        self.bind(pos=self._update, size=self._update, state=state_handler, text=self._update)
        state_handler()
        self._update()

    def _fit_single_line_text(self):
        txt = (self.text or '').strip()
        if not txt:
            self.font_size = self._font_max
            return
        available = max(dp(20), self.width - dp(20))
        chosen = self._font_max
        chosen_w = 0
        for fs in range(int(self._font_max), int(self._font_min) - 1, -1):
            probe = CoreLabel(text=txt, font_size=fs, bold=self.bold)
            probe.refresh()
            tw = (probe.texture.size[0] if probe.texture else 0)
            chosen_w = max(chosen_w, tw)
            if tw <= available:
                chosen = fs
                chosen_w = tw
                break
        self.font_size = chosen
        if self.size_hint_x is None:
            self.width = max(self.width, chosen_w + dp(20))

    def _update(self, *args):
        self.rect.pos, self.rect.size = self.pos, self.size
        self.border_line.rounded_rectangle = (self.x, self.y, self.width, self.height, dp(12))
        self.text_size = (None, None)
        self._fit_single_line_text()

    def _fallback_update_state(self, *args):
        factor = 0.82 if self.state == 'down' else 1.0
        self.bg.rgba = (
            min(1, self.base_color[0] * factor),
            min(1, self.base_color[1] * factor),
            min(1, self.base_color[2] * factor),
            self.base_color[3],
        )

    def _update_state(self, *args):
        factor = 0.82 if self.state == 'down' else 1.0
        self.bg.rgba = (
            min(1, self.base_color[0] * factor),
            min(1, self.base_color[1] * factor),
            min(1, self.base_color[2] * factor),
            self.base_color[3],
        )

class ModernInput(TextInput):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_normal = self.background_active = ""
        self.background_color = (0, 0, 0, 0)
        self.foreground_color = COLOR_TEXT
        self.cursor_color = COLOR_PRIMARY
        self.hint_text_color = (0.7, 0.75, 0.82, 1)
        self.padding = [dp(12), dp(12)]
        with self.canvas.before:
            Color(0.14, 0.17, 0.24, 1)
            self.input_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(10)])
            Color(1, 1, 1, 0.08)
            self.input_border = Line(rounded_rectangle=(self.x, self.y, self.width, self.height, dp(10)), width=1)
        self.bind(pos=self._update_input, size=self._update_input)

    def _update_input(self, *args):
        self.input_rect.pos = self.pos
        self.input_rect.size = self.size
        self.input_border.rounded_rectangle = (self.x, self.y, self.width, self.height, dp(10))

class ColorSafeLabel(Label):
    def __init__(self, bg_color=(1,1,1,1), text_color=(1,1,1,1), **kwargs):
        super().__init__(**kwargs)
        self.color = text_color
        with self.canvas.before:
            Color(*bg_color)
            self.rect = Rectangle(size=self.size, pos=self.pos)
        self.bind(size=self._update, pos=self._update)
    def _update(self, inst, val):
        self.rect.size, self.rect.pos = self.size, self.pos
        self.text_size = (self.width - dp(10), None)

DARK_THEME = {
    "background": (0.07, 0.08, 0.10, 1),
    "card": (0.12, 0.15, 0.20, 1),
    "primary": (0.23, 0.51, 0.96, 1),
    "secondary": (0.21, 0.24, 0.31, 1),
    "danger": (0.80, 0.24, 0.28, 1),
    "text": (0.90, 0.92, 0.95, 1),
    "muted": (0.70, 0.74, 0.80, 1),
}

LIGHT_THEME = {
    "background": (0.96, 0.97, 0.99, 1),
    "card": (1, 1, 1, 1),
    "primary": (0.15, 0.39, 0.92, 1),
    "secondary": (0.86, 0.89, 0.95, 1),
    "danger": (0.82, 0.22, 0.24, 1),
    "text": (0.10, 0.13, 0.18, 1),
    "muted": (0.35, 0.40, 0.48, 1),
}


class AppTheme:
    current = "dark"

    @classmethod
    def palette(cls):
        return DARK_THEME if cls.current == "dark" else LIGHT_THEME


class Card(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.padding = kwargs.get("padding", dp(12))
        self.spacing = kwargs.get("spacing", dp(8))
        with self.canvas.before:
            Color(*AppTheme.palette()["card"])
            self._card_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[dp(12)])
        self.bind(pos=self._update_bg, size=self._update_bg)

    def _update_bg(self, *_):
        self._card_rect.pos = self.pos
        self._card_rect.size = self.size


class PrimaryButton(ModernButton):
    def __init__(self, **kwargs):
        super().__init__(bg_color=AppTheme.palette()["primary"], **kwargs)
        self.height = max(self.height, dp(48))
        self.size_hint_y = None if self.size_hint_y is None else self.size_hint_y


class SecondaryButton(ModernButton):
    def __init__(self, **kwargs):
        super().__init__(bg_color=AppTheme.palette()["secondary"], **kwargs)
        self.height = max(self.height, dp(48))


class DangerButton(ModernButton):
    def __init__(self, **kwargs):
        super().__init__(bg_color=AppTheme.palette()["danger"], **kwargs)
        self.height = max(self.height, dp(48))


class TopBar(BoxLayout):
    def __init__(self, title="", **kwargs):
        super().__init__(orientation="horizontal", size_hint_y=None, height=dp(56), padding=[dp(10), dp(8)], spacing=dp(8), **kwargs)
        with self.canvas.before:
            Color(*COLOR_HEADER)
            self._rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[0, 0, dp(12), dp(12)])
        self.bind(pos=self._upd, size=self._upd)
        self.add_widget(Label(text=title, bold=True, halign="left", valign="middle", color=COLOR_TEXT))

    def _upd(self, *_):
        self._rect.pos = self.pos
        self._rect.size = self.size


class SearchBar(BoxLayout):
    def __init__(self, hint_text="Szukaj...", on_text=None, **kwargs):
        super().__init__(size_hint_y=None, height=dp(54), spacing=dp(8), **kwargs)
        self.input = ModernInput(hint_text=hint_text)
        if on_text:
            self.input.bind(text=on_text)
        self.add_widget(self.input)


class ButtonContainer(ScrollView):
    """Reusable container for action buttons.

    Layout rules:
    - keeps consistent spacing/padding and minimum button size,
    - prevents overlap by forcing explicit button widths/heights,
    - automatically enables scrolling when content overflows.
    """

    def __init__(self, orientation='horizontal', min_button_width=dp(138), min_button_height=dp(48), **kwargs):
        self.orientation = orientation
        self.min_button_width = min_button_width
        self.min_button_height = min_button_height
        kwargs.setdefault('size_hint_y', None)
        kwargs.setdefault('height', dp(64) if orientation == 'horizontal' else dp(200))
        kwargs.setdefault('do_scroll_x', orientation == 'horizontal')
        kwargs.setdefault('do_scroll_y', orientation == 'vertical')
        kwargs.setdefault('bar_width', dp(6))
        super().__init__(**kwargs)

        # Inner layout is sized by minimum_* bindings so overflow is handled by ScrollView.
        if orientation == 'horizontal':
            self.row = BoxLayout(orientation='horizontal', size_hint_x=None, spacing=dp(8), padding=[dp(8), dp(6)])
            self.row.bind(minimum_width=self.row.setter('width'))
        else:
            self.row = GridLayout(cols=1, size_hint_y=None, spacing=dp(8), padding=[dp(6), dp(6)])
            self.row.bind(minimum_height=self.row.setter('height'))

        self.add_widget(self.row)

    def _calc_btn_width(self, widget):
        txt = (getattr(widget, 'text', '') or '').strip()
        probe = CoreLabel(text=txt, font_size=getattr(widget, 'font_size', dp(16)), bold=getattr(widget, 'bold', True))
        probe.refresh()
        tw = probe.texture.size[0] if probe.texture else dp(90)
        return max(self.min_button_width, tw + dp(42))

    def _normalize_button(self, widget):
        widget.size_hint_x = None
        widget.size_hint_y = None
        widget.height = max(getattr(widget, 'height', 0), self.min_button_height)
        if self.orientation == 'horizontal':
            widget.width = max(getattr(widget, 'width', 0), self._calc_btn_width(widget))
        else:
            widget.width = max(getattr(widget, 'width', 0), self.min_button_width)

    def add_action(self, widget):
        self._normalize_button(widget)
        self.row.add_widget(widget)
        if self.orientation == 'horizontal':
            # Keep first actions visible after dynamic updates.
            Clock.schedule_once(lambda dt: setattr(self, 'scroll_x', 0), 0)


class AppActionBar(ButtonContainer):
    """Backward-compatible alias used across existing screens."""

    def __init__(self, **kwargs):
        kwargs.setdefault('orientation', 'horizontal')
        kwargs.setdefault('height', dp(64))
        super().__init__(**kwargs)


class FloatingActionButton(PrimaryButton):
    def __init__(self, **kwargs):
        kwargs.setdefault("text", "+")
        kwargs.setdefault("size_hint", (None, None))
        kwargs.setdefault("size", (dp(58), dp(58)))
        super().__init__(**kwargs)
        self.font_size = '26sp'


class AppLayout(FloatLayout):
    def __init__(self, title="", **kwargs):
        super().__init__(**kwargs)
        self.base = BoxLayout(orientation="vertical", padding=[dp(10), dp(10), dp(10), dp(10)], spacing=dp(8), size_hint=(1, 1))
        self.topbar = TopBar(title=title)
        self.nav_tabs = AppActionBar()
        self.content = BoxLayout(orientation="vertical")
        self.action_bar = AppActionBar()
        self.base.add_widget(self.topbar)
        self.base.add_widget(self.nav_tabs)
        self.base.add_widget(self.content)
        self.base.add_widget(self.action_bar)
        self.add_widget(self.base)
        self.fab = None

    def set_content(self, widget):
        self.content.clear_widgets()
        self.content.add_widget(widget)

    def set_fab(self, on_press):
        if self.fab is not None:
            self.remove_widget(self.fab)
        self.fab = FloatingActionButton(on_press=on_press, pos_hint={"right": 0.97, "y": 0.03})
        self.add_widget(self.fab)

class ClothesSizesScreen(Screen):
    def on_enter(self):
        if not hasattr(self, 'built'):
            self.build_ui()
        self.refresh()

    def build_ui(self):
        try:
            self.clear_widgets()
        except:
            pass
        root = BoxLayout(orientation='vertical')
        top = BoxLayout(size_hint_y=None, height=dp(60), padding=dp(8))
        lbl = Label(text="Rozmiary pracowników", bold=True, size_hint_x=0.7)
        top.add_widget(lbl)
        top.add_widget(ModernButton(text="Dodaj", size_hint_x=0.15, on_press=lambda x: App.get_running_app().form_clothes_size()))
        top.add_widget(ModernButton(text="Wróć", size_hint_x=0.15, on_press=lambda x: setattr(App.get_running_app().sm, 'current', 'clothes')))
        root.add_widget(top)

        sc = ScrollView()
        self.list_layout = GridLayout(cols=1, size_hint_y=None, spacing=dp(6), padding=dp(6))
        self.list_layout.bind(minimum_height=self.list_layout.setter('height'))
        sc.add_widget(self.list_layout)
        root.add_widget(sc)

        self.add_widget(root)
        self.built = True

    def refresh(self):
        self.list_layout.clear_widgets()
        rows = App.get_running_app().conn.execute(
            "SELECT id, name, surname, plant, shirt, hoodie, pants, jacket, shoes FROM clothes_sizes ORDER BY surname"
        ).fetchall()
        for r in rows:
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(285), padding=dp(10), spacing=dp(8))
            with card.canvas.before:
                Color(*COLOR_CARD)
                card_rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(12)])
            card.bind(pos=lambda inst, val, rect=card_rect: setattr(rect, 'pos', val))
            card.bind(size=lambda inst, val, rect=card_rect: setattr(rect, 'size', val))

            full_name = f"{r[1]} {r[2]}"
            head = Label(text=full_name, bold=True, font_size='18sp', halign='left', valign='middle', size_hint_y=None, height=dp(36))
            head.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(8), None)))
            card.add_widget(head)

            details = [
                f"Zakład: {r[3] if str(r[3]).strip() else '-'}",
                f"Rozmiar koszulki: {r[4] if str(r[4]).strip() else '-'}",
                f"Rozmiar bluzy: {r[5] if str(r[5]).strip() else '-'}",
                f"Rozmiar spodni: {r[6] if str(r[6]).strip() else '-'}",
                f"Rozmiar kurtki: {r[7] if str(r[7]).strip() else '-'}",
                f"Rozmiar butów: {r[8] if str(r[8]).strip() else '-'}",
            ]
            for line in details:
                lbl = Label(text=line, halign='left', valign='middle', size_hint_y=None, height=dp(24), color=(0.88, 0.9, 0.96, 1))
                lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(8), None)))
                card.add_widget(lbl)

            btns = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(132), min_button_height=dp(44))
            btns.add_action(ModernButton(text="Edytuj", on_press=lambda x, data=r: App.get_running_app().edit_clothes_size(data)))
            btns.add_action(ModernButton(text="Usuń", bg_color=(0.7,0.1,0.1,1), on_press=lambda x, data=r: App.get_running_app().delete_clothes_size(data[0])))
            card.add_widget(btns)
            self.list_layout.add_widget(card)


class ClothesOrdersScreen(Screen):
    def on_enter(self):
        if not hasattr(self, 'built'):
            self.build_ui()
        self.refresh()

    def build_ui(self):
        try:
            self.clear_widgets()
        except:
            pass
        root = BoxLayout(orientation='vertical', spacing=dp(6))
        header = BoxLayout(size_hint_y=None, height=dp(60), padding=dp(8), spacing=dp(8))
        header.add_widget(Label(text="Zamówienia", bold=True, size_hint_x=0.6))
        header.add_widget(ModernButton(text="Nowe zamówienie", size_hint_x=0.2, on_press=lambda x: App.get_running_app().create_order_ui()))
        header.add_widget(ModernButton(text="Wróć", size_hint_x=0.2, on_press=lambda x: setattr(self.manager, 'current', 'clothes')))
        root.add_widget(header)

        sc = ScrollView()
        self.list_layout = GridLayout(cols=1, size_hint_y=None, spacing=dp(6), padding=dp(6))
        self.list_layout.bind(minimum_height=self.list_layout.setter('height'))
        sc.add_widget(self.list_layout)
        root.add_widget(sc)
        self.add_widget(root)
        self.built = True

    def refresh(self):
        self.list_layout.clear_widgets()
        rows = App.get_running_app().conn.execute("""
        SELECT id,date,plant,status,COALESCE(order_desc,'') FROM clothes_orders ORDER BY id DESC
        """).fetchall()
        for r in rows:
            box = BoxLayout(size_hint_y=None, height=dp(90), padding=dp(6), spacing=dp(8))
            desc = (r[4] or '').strip()
            desc_txt = f"\nOpis: {desc}" if desc else ""
            lbl = Label(text=f"#{r[0]}  {r[1]}  {r[2]}  [{r[3]}]{desc_txt}", size_hint_x=0.55, halign='left', valign='middle')
            lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(12), None)))

            actions = BoxLayout(size_hint_x=0.45, spacing=dp(6))
            actions.add_widget(ModernButton(text="Szczegóły", size_hint_x=None, width=dp(90), on_press=lambda x, i=r[0]: App.get_running_app().clothes_order_details(i)))
            actions.add_widget(ModernButton(text="Zamów", size_hint_x=None, width=dp(90), on_press=lambda x, i=r[0]: App.get_running_app().mark_order_ordered(i)))
            actions.add_widget(ModernButton(text="WYDAJ", size_hint_x=None, width=dp(80), on_press=lambda x, i=r[0]: App.get_running_app().clothes_issue_all(i)))

            box.add_widget(lbl)
            box.add_widget(actions)
            self.list_layout.add_widget(box)

class ClothesStatusScreen(Screen):
    def on_enter(self):
        if not hasattr(self, 'built'):
            self.build_ui()
        self.refresh()

    def build_ui(self):
        try:
            self.clear_widgets()
        except:
            pass
        root = BoxLayout(orientation='vertical')
        sc = ScrollView()
        self.list_layout = GridLayout(cols=1, size_hint_y=None, spacing=dp(6), padding=dp(6))
        self.list_layout.bind(minimum_height=self.list_layout.setter('height'))
        sc.add_widget(self.list_layout)
        root.add_widget(sc)
        self.add_widget(root)
        self.built = True

    def refresh(self):
        self.list_layout.clear_widgets()
        rows = App.get_running_app().conn.execute("""
        SELECT id,date,plant,status FROM clothes_orders ORDER BY id DESC
        """).fetchall()
        for r in rows:
            box = BoxLayout(size_hint_y=None,height=dp(70), padding=dp(6))
            box.add_widget(Label(text=f"Zamówienie #{r[0]}  {r[2]}  {r[3]}"))
            box.add_widget(ModernButton(text="Zmień", size_hint_x=0.25, on_press=lambda x,i=r[0]: App.get_running_app().mark_order_ordered(i)))
            self.list_layout.add_widget(box)

class ClothesReportsScreen(Screen):
    def on_enter(self):
        if not hasattr(self, 'built'):
            self.build_ui()

    def build_ui(self):
        try:
            self.clear_widgets()
        except:
            pass
        root = BoxLayout(orientation='vertical', padding=dp(6), spacing=dp(6))
        header = BoxLayout(size_hint_y=None, height=dp(50))
        header.add_widget(Label(text="Raporty wydanych ubrań", bold=True))
        header.add_widget(ModernButton(text="Export CSV", size_hint_x=None, width=dp(180), on_press=lambda x: App.get_running_app().export_clothes_history_csv()))
        root.add_widget(header)
        self.add_widget(root)
        self.built = True

    def generate(self):
        if SimpleDocTemplate is None:
            App.get_running_app().msg("Brak biblioteki", "Brak reportlab - PDF niedostępny")
            return
        db = App.get_running_app()
        rows = db.conn.execute("""
        SELECT ch.worker_id, w.name, w.surname, ch.item, COUNT(*) as cnt
        FROM clothes_history ch
        LEFT JOIN workers w ON w.id=ch.worker_id
        GROUP BY ch.worker_id, ch.item
        """).fetchall()
        path = Path(db.user_data_dir)/"raport_clothes.pdf"
        c = canvas.Canvas(str(path))
        y = 800
        for r in rows:
            worker = f"{r[1] or ''} {r[2] or ''}".strip()
            txt = f"{worker} {r[3]} {r[4]}"
            c.drawString(50,y,txt)
            y -= 20
            if y < 50:
                c.showPage()
                y = 800
        c.save()
        App.get_running_app().msg("OK", f"Zapisano: {path.name}")
        db.log(f"Generated clothes report: {path}")


class ProUIStyler:
    """Pozostawione dla kompatybilności; brak inwazyjnych modyfikacji runtime UI."""

    def __init__(self):
        self._scan_event = None

    def start(self, root_widget):
        return



class FutureApp(App):
    def build(self):
        Window.clearcolor = AppTheme.palette()["background"]
        if platform == "android":
            from android.permissions import request_permissions, Permission
            request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])
        if not os.path.exists(self.user_data_dir): os.makedirs(self.user_data_dir, exist_ok=True)

        self.full_data, self.filtered_data, self.export_indices = [], [], []
        self.global_attachments, self.queue = [], []
        self.stats = {"ok": 0, "fail": 0, "skip": 0}
        self.idx_name, self.idx_surname, self.idx_pesel = 0, 1, -1
        self.auto_send_mode = False
        self.is_mailing_running = False
        self.mailing_paused = False
        self._log_buffer = []

        self.log_file = Path(self.user_data_dir) / "future_v20.log"
        try:
            self.log_file.touch(exist_ok=True)
        except:
            pass

        self.install_crash_handlers()

        try:
            self.init_db()
        except Exception:
            self.write_crash_report(traceback.format_exc(), "init_db")
        self._screen_initialized = set()

        self.sm = ScreenManager(transition=SlideTransition())
        try:
            self.add_screens()
        except Exception:
            crash_text = traceback.format_exc()
            self.log(f"fatal add_screens error: {crash_text}")
            self.write_crash_report(crash_text, "add_screens")
            self._build_fallback_home()

        self._setup_back_navigation()
        return self.sm

    def _setup_back_navigation(self):
        self._nav_history = []
        self._restoring_nav = False
        Window.bind(on_keyboard=self._on_global_keyboard)
        self.sm.bind(current=lambda *_: self._push_nav_state())
        self._push_nav_state()

    def _bind_clothes_navigation(self):
        if hasattr(self, 'clothes_sm') and not getattr(self, '_clothes_nav_bound', False):
            self.clothes_sm.bind(current=lambda *_: self._push_nav_state())
            self._clothes_nav_bound = True

    def _current_nav_state(self):
        main = self.sm.current if hasattr(self, 'sm') and self.sm else 'home'
        sub = None
        if main == 'clothes' and hasattr(self, 'clothes_sm') and self.clothes_sm:
            sub = self.clothes_sm.current
        return (main, sub)

    def _push_nav_state(self, *_):
        if getattr(self, '_restoring_nav', False):
            return
        state = self._current_nav_state()
        hist = getattr(self, '_nav_history', None)
        if hist is None:
            self._nav_history = [state]
            return
        if not hist or hist[-1] != state:
            hist.append(state)
        if len(hist) > 150:
            del hist[0:len(hist)-150]

    def _apply_nav_state(self, state):
        target_main, target_sub = state
        if target_main not in getattr(self, 'sc_ref', {}):
            target_main = 'home'
        self.ensure_screen_ui(target_main)
        self.sm.current = target_main
        if target_main == 'clothes':
            self._bind_clothes_navigation()
            if target_sub and hasattr(self, 'clothes_sm') and self.clothes_sm.has_screen(target_sub):
                self.clothes_sm.current = target_sub

    def go_back(self):
        hist = getattr(self, '_nav_history', [])
        if len(hist) <= 1:
            return False
        self._restoring_nav = True
        try:
            hist.pop()
            self._apply_nav_state(hist[-1])
        finally:
            self._restoring_nav = False
        return True

    def _on_global_keyboard(self, window, key, scancode, codepoint, modifiers):
        if key in (27, 1001):
            if self.go_back():
                return True
            if getattr(self, 'sm', None) and self.sm.current != 'home':
                self.ensure_screen_ui('home')
                self.sm.current = 'home'
                self._push_nav_state()
                return True
            return False
        return False

    def switch_theme(self, mode):
        if mode not in ("dark", "light"):
            return
        AppTheme.current = mode
        pal = AppTheme.palette()
        Window.clearcolor = pal["background"]
        self.setup_ui_all()
        for screen_name in ("contacts", "cars", "pracownicy", "zaklady", "settings", "paski"):
            try:
                self._screen_initialized.discard(screen_name)
                self.ensure_screen_ui(screen_name)
            except Exception:
                pass

    def _documents_dir(self):
        if platform == "android":
            return Path("/storage/emulated/0/Documents")
        return Path.home() / "Documents"

    def write_crash_report(self, details, where="runtime"):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            content = (
                f"FUTURE ULTIMATE CRASH REPORT\n"
                f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Where: {where}\n"
                f"Platform: {platform}\n"
                f"\n{details}\n"
            )
            targets = [Path(self.user_data_dir), self._documents_dir()]
            for target in targets:
                try:
                    target.mkdir(parents=True, exist_ok=True)
                    report = target / f"future_crash_{ts}.txt"
                    with open(report, "w", encoding="utf-8") as f:
                        f.write(content)
                except Exception:
                    pass
            self.log(f"Crash report generated ({where})")
        except Exception:
            pass

    def install_crash_handlers(self):
        def _handle(exc_type, exc, tb):
            text = ''.join(traceback.format_exception(exc_type, exc, tb))
            self.write_crash_report(text, "sys.excepthook")
        sys.excepthook = _handle

        def _thread_handle(args):
            text = ''.join(traceback.format_exception(args.exc_type, args.exc_value, args.exc_traceback))
            self.write_crash_report(text, "threading.excepthook")
        try:
            threading.excepthook = _thread_handle
        except Exception:
            pass

    def _build_fallback_home(self):
        try:
            self.sm.clear_widgets()
        except Exception:
            pass
        sc = Screen(name="home")
        root = BoxLayout(orientation="vertical", padding=dp(16), spacing=dp(10))
        root.add_widget(Label(text="FUTURE ULTIMATE v20", bold=True, font_size='26sp', color=COLOR_PRIMARY))
        root.add_widget(Label(text="Uruchomiono tryb awaryjny. Sprawdź plik crash .txt w Documents lub future_v20.log.", halign='center'))
        root.add_widget(ModernButton(text="Pokaż logi", on_press=lambda x: self.show_logs()))
        root.add_widget(ModernButton(text="Zamknij", bg_color=(0.65,0.18,0.2,1), on_press=lambda x: self.stop()))
        sc.add_widget(root)
        self.sm.add_widget(sc)
        self.sm.current = "home"

    def log(self, txt):
        try:
            t = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            line = f"[{t}] {txt}\n"
            self._log_buffer.append(line)
            if len(self._log_buffer) > 200:
                self._log_buffer = self._log_buffer[-200:]
            with open(self.log_file, "a", encoding="utf-8") as f:
                f.write(line)
        except:
            pass

    def _add_column_if_missing(self, table, column, ctype='TEXT'):
        try:
            cols = [r[1] for r in self.conn.execute(f"PRAGMA table_info({table})").fetchall()]
            if column not in cols:
                self.conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ctype}")
                self.conn.commit()
        except Exception:
            pass

    def patch_contact_extra_fields(self):
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN workplace TEXT")
        except:
            pass
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN apartment TEXT")
        except:
            pass
        self.conn.commit()

    def patch_contacts_database(self):
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN plant TEXT")
        except:
            pass
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN hire_date TEXT")
        except:
            pass
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN clothes_size TEXT")
        except:
            pass
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN shoes_size TEXT")
        except:
            pass
        try:
            self.conn.execute("ALTER TABLE contacts ADD COLUMN notes TEXT")
        except:
            pass
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS clothes_history(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            worker_id INTEGER,
            name TEXT,
            surname TEXT,
            item TEXT,
            size TEXT,
            date TEXT
        )
        """)
        self.conn.commit()

    def init_cars_db(self):
        """Tworzy i uzupełnia tabelę modułu samochodów."""
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS cars(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            registration TEXT,
            driver TEXT,
            mileage INTEGER DEFAULT 0,
            service_interval INTEGER DEFAULT 15000,
            last_service INTEGER DEFAULT 0
        )
        """)
        self.conn.commit()

    def ensure_extended_tables(self):
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS plants(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        city TEXT,
        address TEXT,
        contact_phone TEXT,
        notes TEXT
        )
        """)
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS fleet_cars(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plate TEXT UNIQUE,
        brand TEXT,
        model TEXT,
        plant TEXT,
        mileage INTEGER DEFAULT 0,
        status TEXT,
        driver TEXT,
        notes TEXT
        )
        """)
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS workers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        surname TEXT,
        plant TEXT
        )
        """)
        self._add_column_if_missing('workers', 'phone', 'TEXT')
        self._add_column_if_missing('workers', 'position', 'TEXT')
        self._add_column_if_missing('workers', 'hire_date', 'TEXT')
        self.init_cars_db()
        self.conn.commit()

    def init_db(self):
        db_p = Path(self.user_data_dir) / "future_v20.db"
        self.conn = sqlite3.connect(str(db_p), check_same_thread=False)
        self.conn.execute("CREATE TABLE IF NOT EXISTS contacts (name TEXT, surname TEXT, email TEXT, pesel TEXT, phone TEXT, PRIMARY KEY(name, surname))")
        self.conn.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, val TEXT)")
        self.conn.execute("CREATE TABLE IF NOT EXISTS reports (id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT, ok INTEGER, fail INTEGER, skip INTEGER, auto INTEGER, details TEXT)")
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS plants(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        city TEXT,
        address TEXT,
        contact_phone TEXT,
        notes TEXT
        )
        """)
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS fleet_cars(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plate TEXT UNIQUE,
        brand TEXT,
        model TEXT,
        plant TEXT,
        mileage INTEGER DEFAULT 0,
        status TEXT,
        driver TEXT,
        notes TEXT
        )
        """)
        self.init_cars_db()
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS clothes_sizes(
        id INTEGER PRIMARY KEY,
        name TEXT,
        surname TEXT,
        plant TEXT,
        shirt TEXT,
        hoodie TEXT,
        pants TEXT,
        jacket TEXT,
        shoes TEXT
        )
        """)
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS clothes_orders(
        id INTEGER PRIMARY KEY,
        date TEXT,
        plant TEXT,
        status TEXT,
        order_desc TEXT
        )
        """)
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS clothes_order_items(
        id INTEGER PRIMARY KEY,
        order_id INTEGER,
        worker_id INTEGER,
        name TEXT,
        surname TEXT,
        item TEXT,
        size TEXT,
        qty INTEGER,
        issued INTEGER DEFAULT 0
        )
        """)
        self.conn.execute("""
        CREATE TABLE IF NOT EXISTS clothes_issued(
        id INTEGER PRIMARY KEY,
        name TEXT,
        surname TEXT,
        item TEXT,
        size TEXT,
        qty INTEGER,
        date TEXT
        )
        """)
        self.ensure_extended_tables()
        self.conn.commit()
        try:
            self.patch_contact_extra_fields()
        except:
            pass
        try:
            self.patch_contacts_database()
        except:
            pass
        try:
            self.clothes_init()
        except:
            pass
        try:
            self.sync_all_contact_links()
        except Exception:
            self.log(f"sync_all_contact_links error: {traceback.format_exc()}")

    def clothes_init(self):
        c=self.conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS workers(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        surname TEXT,
        plant TEXT
        )
        """)
        c.execute("""
        CREATE TABLE IF NOT EXISTS worker_sizes(
        worker_id INTEGER,
        shirt TEXT,
        pants TEXT,
        shoes TEXT,
        jacket TEXT
        )
        """)
        self.conn.commit()
        self._add_column_if_missing('clothes_history', 'worker_id', 'INTEGER')
        self._add_column_if_missing('clothes_history', 'name', 'TEXT')
        self._add_column_if_missing('clothes_history', 'surname', 'TEXT')
        self._add_column_if_missing('clothes_history', 'item', 'TEXT')
        self._add_column_if_missing('clothes_history', 'size', 'TEXT')
        self._add_column_if_missing('clothes_history', 'date', 'TEXT')
        self._add_column_if_missing('clothes_order_items', 'worker_id', 'INTEGER')
        self._add_column_if_missing('clothes_order_items', 'item', 'TEXT')
        self._add_column_if_missing('clothes_order_items', 'qty', 'INTEGER')
        self._add_column_if_missing('clothes_order_items', 'issued', 'INTEGER')
        self._add_column_if_missing('clothes_orders', 'order_desc', 'TEXT')
        self._add_column_if_missing('workers', 'phone', 'TEXT')
        self._add_column_if_missing('workers', 'position', 'TEXT')
        self._add_column_if_missing('workers', 'hire_date', 'TEXT')

    def clothes_import_excel(self,path):
        if pd is None:
            self.msg("Błąd","Brak biblioteki pandas - import niemożliwy")
            return
        df=pd.read_excel(path)
        col_map = {str(col).lower(): col for col in df.columns}
        name_col = None
        surname_col = None
        plant_col = None
        for key in col_map:
            if "imi" in key or "name" in key:
                name_col = col_map[key]
            if "naz" in key or "surname" in key:
                surname_col = col_map[key]
            if "zak" in key or "plant" in key:
                plant_col = col_map[key]
        if not name_col or not surname_col:
            self.msg("Błąd","Nie znaleziono kolumn imię/nazwisko")
            return
        c=self.conn.cursor()
        for _,row in df.iterrows():
            name=row[name_col]
            surname=row[surname_col]
            plant=row[plant_col] if plant_col else ""
            try:
                c.execute("""
                INSERT INTO workers(name,surname,plant)
                VALUES(?,?,?)
                """,(str(name).strip(),str(surname).strip(),str(plant).strip()))
            except:
                pass
        self.conn.commit()
        self.msg("OK","Import zakończony")

    def clothes_edit_sizes(self,worker_id):
        root=BoxLayout(orientation="vertical",padding=dp(10),spacing=dp(6))
        shirt=TextInput(hint_text="Koszulka")
        pants=TextInput(hint_text="Spodnie")
        shoes=TextInput(hint_text="Buty")
        jacket=TextInput(hint_text="Kurtka")
        cur = self.conn.cursor()
        existing = cur.execute("SELECT shirt,pants,shoes,jacket FROM worker_sizes WHERE worker_id=?", (worker_id,)).fetchone()
        if existing:
            shirt.text, pants.text, shoes.text, jacket.text = existing[0] or "", existing[1] or "", existing[2] or "", existing[3] or ""
        root.add_widget(shirt)
        root.add_widget(pants)
        root.add_widget(shoes)
        root.add_widget(jacket)
        def save(_):
            try:
                self.conn.execute("DELETE FROM worker_sizes WHERE worker_id=?", (worker_id,))
            except:
                pass
            self.conn.execute("""
            INSERT INTO worker_sizes(worker_id,shirt,pants,shoes,jacket)
            VALUES(?,?,?,?,?)
            """,(worker_id,shirt.text,pants.text,shoes.text,jacket.text))
            self.conn.commit()
            self.msg("OK","Rozmiary zapisane")
            px.dismiss()
        root.add_widget(ModernButton(text="ZAPISZ",on_press=save))
        px = Popup(title="ROZMIARY",content=root,size_hint=(0.8,0.8))
        px.open()

    def clothes_select_workers(self):
        root=BoxLayout(orientation="vertical",padding=dp(10),spacing=dp(6))
        plant_filter=TextInput(hint_text="Zakład")
        root.add_widget(plant_filter)
        cur=self.conn.cursor()
        workers=[]
        grid=GridLayout(cols=1,size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))
        rows=cur.execute("""
        SELECT id,name,surname,plant
        FROM workers
        ORDER BY surname
        """).fetchall()
        for r in rows:
            wid=r[0]
            label=f"{r[1]} {r[2]} ({r[3]})"
            cb=CheckBox()
            row=BoxLayout(size_hint_y=None,height=dp(36))
            row.add_widget(Label(text=label))
            row.add_widget(cb)
            grid.add_widget(row)
            workers.append((wid,r[3],cb))
        scroll=ScrollView(size_hint=(1,1))
        scroll.add_widget(grid)
        root.add_widget(scroll)
        def select_plant(_):
            plant=plant_filter.text.strip().lower()
            for wid,p,cb in workers:
                cb.active=(p and p.strip().lower()==plant)
        root.add_widget(ModernButton(text="WYBIERZ ZAKŁAD",on_press=select_plant))
        p = Popup(title="WYBÓR PRACOWNIKÓW",content=root,size_hint=(0.9,0.9))
        p.open()
        return workers, p

    def clothes_create_order(self, worker_ids, items, plant):
        c=self.conn.cursor()
        c.execute("""
        INSERT INTO clothes_orders(date,plant,status)
        VALUES(?,?,?)
        """,(datetime.now().strftime("%Y-%m-%d"), plant or "Zakład", "Do zamówienia"))
        order_id=c.lastrowid
        for wid in worker_ids:
            for item in items:
                c.execute("""
                INSERT INTO clothes_order_items(order_id, worker_id, name, surname, item, qty, issued)
                VALUES(?,?,?,?,?,?,?)
                """,(order_id, wid, "", "", item.get('name',''), item.get('qty',1), 0))
        self.conn.commit()
        self.msg("OK","Zamówienie utworzone")
        return order_id

    def clothes_count_sizes(self,order_id):
        cur=self.conn.cursor()
        rows=cur.execute("""
        SELECT worker_id,item,qty
        FROM clothes_order_items
        WHERE order_id=?
        """,(order_id,)).fetchall()
        summary=defaultdict(int)
        for wid,item,qty in rows:
            size=cur.execute("""
            SELECT shirt,pants,shoes,jacket
            FROM worker_sizes
            WHERE worker_id=?
            """,(wid,)).fetchone()
            if not size:
                summary[f"{item} (brak rozmiaru)"] += (qty or 1)
                continue
            item_l = (item or "").lower()
            if "kosz" in item_l or "koszulka" in item_l:
                s = size[0]
            elif "spod" in item_l or "spodnie" in item_l:
                s = size[1]
            elif "but" in item_l or "buty" in item_l:
                s = size[2]
            else:
                s = size[3]
            summary[f"{item} {s}"] += (qty or 1)
        return summary

    def clothes_order_pdf(self,order_id):
        if SimpleDocTemplate is None:
            self.msg("PDF","Brak reportlab.platypus - PDF niedostępny")
            return
        summary=self.clothes_count_sizes(order_id)
        styles=getSampleStyleSheet()
        elements=[]
        elements.append(Paragraph("ZAMÓWIENIE UBRAŃ",styles['Title']))
        elements.append(Spacer(1,20))
        data=[["Pozycja","Ilość"]]
        for k,v in summary.items():
            data.append([k,v])
        pdf=SimpleDocTemplate(f"zamowienie_{order_id}.pdf",pagesize=A4)
        pdf.build(elements+[Table(data)])
        self.msg("PDF","PDF zamówienia zapisany")

    def clothes_issue_pdf(self,order_id):
        if SimpleDocTemplate is None:
            self.msg("PDF","Brak reportlab.platypus - PDF niedostępny")
            return
        cur=self.conn.cursor()
        rows=cur.execute("""
        SELECT coi.id, w.name, w.surname, coi.item, coi.qty
        FROM clothes_order_items coi
        LEFT JOIN workers w ON w.id=coi.worker_id
        WHERE coi.order_id=?
        """,(order_id,)).fetchall()
        styles=getSampleStyleSheet()
        elements=[]
        elements.append(Paragraph("LISTA WYDANIA UBRAŃ",styles['Title']))
        elements.append(Spacer(1,20))
        data=[["Pracownik","Ubranie","Ilość","Podpis"]]
        for r in rows:
            worker = f"{r[1] or ''} {r[2] or ''}".strip()
            data.append([worker, r[3], r[4] or 1, ""])
        pdf=SimpleDocTemplate(f"wydanie_{order_id}.pdf",pagesize=A4)
        pdf.build(elements+[Table(data)])
        self.msg("PDF","PDF wydania zapisany")

    def _order_status(self, order_id):
        row = self.conn.execute("SELECT COALESCE(status,'') FROM clothes_orders WHERE id=?", (order_id,)).fetchone()
        return (row[0] if row else '') or ''

    def _can_issue_order(self, order_id):
        st = self._order_status(order_id).strip().lower()
        if st != 'zamówione':
            self.msg("Info", "Najpierw kliknij 'Zamów' (status musi być: Zamówione)")
            return False
        return True

    def _refresh_order_issue_status(self, order_id):
        row = self.conn.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN COALESCE(issued,0)=1 THEN 1 ELSE 0 END), 0),
            COUNT(*)
        FROM clothes_order_items
        WHERE order_id=?
        """, (order_id,)).fetchone()
        issued_cnt = int((row[0] if row else 0) or 0)
        total_cnt = int((row[1] if row else 0) or 0)
        if total_cnt <= 0:
            return
        if issued_cnt >= total_cnt:
            status = 'Wydane'
        elif issued_cnt > 0:
            status = 'Częściowo wydane'
        else:
            return
        self.conn.execute("UPDATE clothes_orders SET status=? WHERE id=?", (status, order_id))

    def clothes_issue_all(self,order_id):
        if not self._can_issue_order(order_id):
            return
        cur=self.conn.cursor()
        rows=cur.execute("""
        SELECT coi.id, coi.worker_id, COALESCE(w.name, coi.name, ''), COALESCE(w.surname, coi.surname, ''), coi.item, COALESCE(coi.size, ''), COALESCE(coi.qty, 1)
        FROM clothes_order_items coi
        LEFT JOIN workers w ON w.id=coi.worker_id
        WHERE coi.order_id=? AND COALESCE(coi.issued,0)=0
        """,(order_id,)).fetchall()
        if not rows:
            self.msg("Info", "Brak niewydanych pozycji")
            return
        for r in rows:
            coi_id, wid, name, surname, item, size, _qty = r
            cur.execute("""
            INSERT INTO clothes_history(worker_id, name, surname, item, size, date)
            VALUES(?,?,?,?,?,?)
            """,(wid, name, surname, item, size, datetime.now().strftime("%Y-%m-%d")))
            cur.execute("UPDATE clothes_order_items SET issued=1 WHERE id=?", (coi_id,))
        self._refresh_order_issue_status(order_id)
        self.conn.commit()
        self.msg("OK","Ubrania wydane")

    def clothes_issue_partial(self,order_id):
        if not self._can_issue_order(order_id):
            return
        root=BoxLayout(orientation="vertical",padding=dp(10),spacing=dp(6))
        cur=self.conn.cursor()
        grid=GridLayout(cols=1,size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))
        items=[]
        rows=cur.execute("""
        SELECT c.id,
               COALESCE(w.name, c.name, ''),
               COALESCE(w.surname, c.surname, ''),
               c.item,
               COALESCE(c.size, ''),
               COALESCE(c.qty,1),
               COALESCE(c.issued,0)
        FROM clothes_order_items c
        LEFT JOIN workers w ON w.id=c.worker_id
        WHERE order_id=?
        """,(order_id,)).fetchall()
        any_available = False
        for r in rows:
            cid=r[0]
            already_issued = int(r[6] or 0) == 1
            if not already_issued:
                any_available = True
            label=f"{r[1] or ''} {r[2] or ''} - {r[3]} {r[4] or '-'} x{r[5] or 1}{' (wydane)' if already_issued else ''}"
            cb=CheckBox()
            cb.disabled = already_issued
            row=BoxLayout(size_hint_y=None,height=dp(36))
            lbl = Label(text=label, halign='left', valign='middle')
            lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(12), None)))
            row.add_widget(lbl)
            row.add_widget(cb)
            grid.add_widget(row)
            items.append((cid,cb,already_issued))
        scroll=ScrollView(size_hint=(1,1))
        scroll.add_widget(grid)
        root.add_widget(scroll)

        if not any_available:
            root.add_widget(Label(text="Brak pozycji do wydania", size_hint_y=None, height=dp(32)))

        def save(_):
            changed = 0
            for cid,cb,already_issued in items:
                if already_issued:
                    continue
                if cb.active:
                    cur.execute("""
                    INSERT INTO clothes_history(worker_id, name, surname, item, size, date)
                    SELECT coi.worker_id, COALESCE(w.name, coi.name, ''), COALESCE(w.surname, coi.surname, ''), coi.item, COALESCE(coi.size, ''), ?
                    FROM clothes_order_items coi
                    LEFT JOIN workers w ON w.id=coi.worker_id
                    WHERE coi.id=?
                    """,(datetime.now().strftime("%Y-%m-%d"),cid))
                    cur.execute("UPDATE clothes_order_items SET issued=1 WHERE id=?", (cid,))
                    changed += 1
            if changed <= 0:
                self.msg("Info", "Nie zaznaczono pozycji do wydania")
                return
            self._refresh_order_issue_status(order_id)
            self.conn.commit()
            self.msg("OK","Wydanie zapisane")
            px.dismiss()
            self.clothes_order_details(order_id)
        root.add_widget(ModernButton(text="WYDAJ ZAZNACZONE",on_press=save))
        px = Popup(title="WYDANIE CZĘŚCIOWE",content=root,size_hint=(0.9,0.9))
        px.open()

    def clothes_worker_year_stats(self,worker_id,year=None):
        if not year:
            year=datetime.now().year
        root=BoxLayout(orientation="vertical",padding=dp(10),spacing=dp(6))
        cur=self.conn.cursor()
        worker=cur.execute("""
        SELECT name,surname,plant
        FROM workers
        WHERE id=?
        """,(worker_id,)).fetchone()
        root.add_widget(Label(
            text=f"{worker[0]} {worker[1]} ({worker[2]}) - {year}",
            size_hint_y=None,
            height=dp(50)
        ))
        clothes=["koszulka","spodnie","buty","kurtka"]
        grid=GridLayout(cols=1,size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))
        for item in clothes:
            count=cur.execute("""
            SELECT COUNT(*)
            FROM clothes_history
            WHERE worker_id=? AND item LIKE ? AND strftime('%Y',date)=?
            """,(worker_id,'%'+item+'%',str(year))).fetchone()[0]
            last=cur.execute("""
            SELECT date
            FROM clothes_history
            WHERE worker_id=? AND item LIKE ?
            ORDER BY date DESC
            LIMIT 1
            """,(worker_id,'%'+item+'%')).fetchone()
            last_date=last[0] if last else "-"
            row=BoxLayout(size_hint_y=None,height=dp(40))
            row.add_widget(Label(
                text=f"{item}   w {year}: {count}   ostatnie: {last_date}"
            ))
            grid.add_widget(row)
        scroll=ScrollView()
        scroll.add_widget(grid)
        root.add_widget(scroll)
        Popup(
            title="Statystyki ubrań",
            content=root,
            size_hint=(0.85,0.85)
        ).open()

    def clothes_stats_panel(self):
        root=BoxLayout(orientation="vertical",padding=dp(10),spacing=dp(6))
        year_input=TextInput(
            text=str(datetime.now().year),
            hint_text="Rok",
            size_hint_y=None,
            height=dp(40)
        )
        root.add_widget(year_input)
        grid=GridLayout(cols=1,size_hint_y=None)
        grid.bind(minimum_height=grid.setter("height"))
        cur=self.conn.cursor()
        rows=cur.execute("""
        SELECT id,name,surname,plant
        FROM workers
        ORDER BY surname
        """).fetchall()
        for r in rows:
            wid=r[0]
            btn=ModernButton(
                text=f"{r[1]} {r[2]} ({r[3]})",
                size_hint_y=None,
                height=dp(40),
                on_press=lambda x,wid=wid:self.clothes_worker_year_stats(
                    wid,
                    int(year_input.text)
                )
            )
            grid.add_widget(btn)
        scroll=ScrollView()
        scroll.add_widget(grid)
        root.add_widget(scroll)
        Popup(
            title="Statystyki roczne ubrań",
            content=root,
            size_hint=(0.9,0.9)
        ).open()

    def add_screens(self):
        names = ["home", "table", "email", "smtp", "tmpl", "contacts", "report", "cars", "clothes", "paski", "pracownicy", "zaklady", "settings"]
        self.sc_ref = {name: Screen(name=name) for name in names}
        self.setup_ui_all()
        for s in self.sc_ref.values():
            self.sm.add_widget(s)
            if s.name != "home":
                s.bind(on_pre_enter=lambda inst, *a: self.ensure_screen_ui(inst.name))
        if "clothes" in self.sc_ref:
            self.sc_ref["clothes"].bind(on_enter=lambda inst, *a: self._on_main_clothes_enter())

    def _on_main_clothes_enter(self):
        try:
            if hasattr(self, 'clothes_sm'):
                self.clothes_sm.current = 'sizes'
                scr = self.clothes_sm.get_screen('sizes')
                if hasattr(scr, 'build_ui'):
                    scr.build_ui()
                if hasattr(scr, 'refresh'):
                    scr.refresh()
        except:
            pass

    def _safe_setup(self, name, fn):
        try:
            fn()
            return True
        except Exception:
            try:
                self.log(f"setup error [{name}]: {traceback.format_exc()}")
            except Exception:
                pass
            return False

    def ensure_screen_ui(self, name):
        if not hasattr(self, '_screen_initialized'):
            self._screen_initialized = set()
        if name in self._screen_initialized:
            return
        mapping = {
            "table": self.setup_table_ui,
            "email": self.setup_email_ui,
            "smtp": self.setup_smtp_ui,
            "tmpl": self.setup_tmpl_ui,
            "contacts": self.setup_contacts_ui,
            "report": self.setup_report_ui,
            "cars": self.setup_cars_ui,
            "paski": self.setup_paski_ui,
            "pracownicy": self.setup_pracownicy_ui,
            "zaklady": self.setup_zaklady_ui,
            "settings": self.setup_settings_ui,
            "clothes": self.setup_clothes_container,
        }
        fn = mapping.get(name)
        if fn is None:
            return
        if self._safe_setup(name, fn):
            self._screen_initialized.add(name)

    def setup_ui_all(self):
        self.sc_ref["home"].clear_widgets()
        layout = AppLayout(title="FUTURE ULTIMATE v20")
        layout.nav_tabs.add_action(SecondaryButton(text="Dark", on_press=lambda x: self.switch_theme("dark")))
        layout.nav_tabs.add_action(SecondaryButton(text="Light", on_press=lambda x: self.switch_theme("light")))

        content = BoxLayout(orientation="vertical", spacing=dp(10), padding=[0, dp(6), 0, 0])
        content.add_widget(Label(text="Panel główny aplikacji", font_size='15sp', color=(0.72, 0.78, 0.9, 1), size_hint_y=None, height=dp(26)))
        sv = ScrollView(size_hint=(1, 1))
        grid = GridLayout(cols=2, spacing=dp(12), padding=dp(6), size_hint_y=None)
        grid.bind(minimum_height=grid.setter('height'))
        btn_props = dict(size_hint_y=None, height=dp(86))
        grid.add_widget(PrimaryButton(text="Kontakty", on_press=lambda x: [self.ensure_screen_ui("contacts"), self.refresh_contacts_list(), setattr(self.sm, 'current', 'contacts')], **btn_props))
        grid.add_widget(PrimaryButton(text="Samochody", on_press=lambda x: setattr(self.sm, 'current', 'cars'), **btn_props))
        grid.add_widget(PrimaryButton(text="Ubranie robocze", on_press=lambda x: setattr(self.sm, 'current', 'clothes'), **btn_props))
        grid.add_widget(PrimaryButton(text="Paski", on_press=lambda x: setattr(self.sm, 'current', 'paski'), **btn_props))
        grid.add_widget(PrimaryButton(text="Pracownicy", on_press=lambda x: setattr(self.sm, 'current', 'pracownicy'), **btn_props))
        grid.add_widget(PrimaryButton(text="Zakłady", on_press=lambda x: setattr(self.sm, 'current', 'zaklady'), **btn_props))
        grid.add_widget(SecondaryButton(text="Ustawienia", on_press=lambda x: setattr(self.sm, 'current', 'settings'), **btn_props))
        grid.add_widget(DangerButton(text="Wyjście", on_press=lambda x: App.get_running_app().stop(), **btn_props))
        sv.add_widget(grid)
        content.add_widget(sv)
        layout.set_content(content)
        self.sc_ref["home"].add_widget(layout)

    def setup_table_ui(self):
        self.sc_ref["table"].clear_widgets()
        shell = AppLayout(title="Podgląd i eksport")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm, 'current', 'paski')))
        shell.nav_tabs.add_action(PrimaryButton(text="Kolumny", on_press=self.popup_columns, size_hint_x=None, width=dp(150)))

        root = BoxLayout(orientation="vertical", spacing=dp(8))
        self.ti_tab_search = ModernInput(hint_text="Szukaj w tabeli...")
        self.ti_tab_search.bind(text=self.filter_table)
        root.add_widget(self.ti_tab_search)

        hs = ScrollView(size_hint_y=None, height=dp(58), do_scroll_y=False)
        self.table_header_layout = GridLayout(rows=1, size_hint=(None, None), height=dp(58))
        hs.add_widget(self.table_header_layout)

        ds = ScrollView(do_scroll_x=True, do_scroll_y=True)
        self.table_content_layout = GridLayout(size_hint=(None, None), spacing=dp(2))
        self.table_content_layout.bind(minimum_height=self.table_content_layout.setter('height'), minimum_width=self.table_content_layout.setter('width'))
        ds.add_widget(self.table_content_layout)
        ds.bind(scroll_x=lambda inst, val: setattr(hs, 'scroll_x', val))

        root.add_widget(hs)
        root.add_widget(ds)
        shell.set_content(root)
        self.sc_ref["table"].add_widget(shell)

    def refresh_table(self):
        self.table_content_layout.clear_widgets()
        self.table_header_layout.clear_widgets()
        if not self.filtered_data:
            return
        w_cell, w_act, h = dp(170), dp(220), dp(55)
        headers = [self.full_data[0][i] for i in self.export_indices]

        total_w = (len(headers) * w_cell) + w_act
        self.table_header_layout.cols = self.table_content_layout.cols = len(headers) + 1
        self.table_header_layout.width = self.table_content_layout.width = total_w

        for head in headers:
            self.table_header_layout.add_widget(ColorSafeLabel(text=str(head), bg_color=COLOR_HEADER, bold=True, size=(w_cell, h), size_hint=(None,None), text_color=(0,0,0,1)))
        self.table_header_layout.add_widget(ColorSafeLabel(text="AKCJE", bg_color=COLOR_HEADER, bold=True, size=(w_act, h), size_hint=(None,None), text_color=(0,0,0,1)))

        for r_idx, row in enumerate(self.filtered_data[1:]):
            row_bg = COLOR_ROW_A if r_idx % 2 == 0 else COLOR_ROW_B
            for c_idx in self.export_indices:
                val = str(row[c_idx]) if c_idx < len(row) and str(row[c_idx]).strip() != "" else "0"
                self.table_content_layout.add_widget(ColorSafeLabel(text=val, bg_color=row_bg, size=(w_cell, h), size_hint=(None,None)))

            act_box = BoxLayout(size=(w_act, h), size_hint=(None,None), spacing=dp(4), padding=dp(4))
            act_box.add_widget(Button(text="ZAPISZ", on_press=lambda x, r=row: self.export_single_row(r), background_color=(0.2, 0.6, 0.2, 1)))
            act_box.add_widget(Button(text="WYŚLIJ", on_press=lambda x, r=row: self.send_individual_from_table(r), background_color=(0.1, 0.5, 0.9, 1)))
            self.table_content_layout.add_widget(act_box)

    def setup_clothes_container(self):
        self.sc_ref["clothes"].clear_widgets()
        shell = AppLayout(title="Ubranie robocze")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm, 'current', 'home')))

        tabs = AppActionBar()
        btn_w = dp(170)
        tabs.add_action(PrimaryButton(text="Rozmiary", size_hint_x=None, width=btn_w, on_press=lambda x: setattr(self.clothes_sm, 'current', 'sizes')))
        tabs.add_action(PrimaryButton(text="Zamówienia", size_hint_x=None, width=btn_w, on_press=lambda x: setattr(self.clothes_sm, 'current', 'orders')))
        tabs.add_action(PrimaryButton(text="Raporty", size_hint_x=None, width=btn_w, on_press=lambda x: setattr(self.clothes_sm, 'current', 'reports')))

        self.clothes_sm = ScreenManager(transition=SlideTransition())
        self.clothes_sm.add_widget(ClothesSizesScreen(name='sizes'))
        self.clothes_sm.add_widget(ClothesOrdersScreen(name='orders'))
        self.clothes_sm.add_widget(ClothesReportsScreen(name='reports'))
        self.clothes_sm.current = 'sizes'
        self._clothes_nav_bound = False
        self._bind_clothes_navigation()

        body = BoxLayout(orientation='vertical', spacing=dp(8))
        body.add_widget(tabs)
        body.add_widget(self.clothes_sm)
        shell.set_content(body)
        self.sc_ref["clothes"].add_widget(shell)
        self._push_nav_state()

        try:
            scr = self.clothes_sm.get_screen('sizes')
            if hasattr(scr, 'build_ui'):
                scr.build_ui()
            if hasattr(scr, 'refresh'):
                scr.refresh()
        except:
            pass

    def _clothes_fetch_workers_for_order(self):
        rows = self.conn.execute("""
        SELECT w.id, w.name, w.surname, COALESCE(NULLIF(w.plant,''), cs.plant, '') AS plant,
               cs.shirt, cs.hoodie, cs.pants, cs.jacket, cs.shoes
        FROM workers w
        LEFT JOIN clothes_sizes cs
          ON lower(cs.name)=lower(w.name) AND lower(cs.surname)=lower(w.surname)
        ORDER BY w.surname, w.name
        """).fetchall()
        out = []
        for r in rows:
            out.append({
                'id': r[0], 'name': r[1] or '', 'surname': r[2] or '', 'plant': r[3] or '',
                'sizes': {'Koszulka': r[4] or '', 'Bluza': r[5] or '', 'Spodnie': r[6] or '', 'Kurtka': r[7] or '', 'Buty': r[8] or ''}
            })
        return out

    def _collect_order_entries(self, selected_workers, worker_forms):
        entries = []
        for w in selected_workers:
            frm = worker_forms.get(w['id'])
            if not frm or not frm['use'].active:
                continue
            for item_name, qty_ti in frm['qty'].items():
                try:
                    qty = int(qty_ti.text.strip() or '0')
                except Exception:
                    qty = 0
                if qty <= 0:
                    continue
                size = w['sizes'].get(item_name, '')
                entries.append({
                    'worker_id': w['id'],
                    'name': w['name'],
                    'surname': w['surname'],
                    'item': item_name,
                    'size': size,
                    'qty': qty,
                })
        return entries

    def _save_clothes_order_entries(self, plant, order_desc, entries):
        cur = self.conn.cursor()
        cur.execute("INSERT INTO clothes_orders(date,plant,status,order_desc) VALUES(?,?,?,?)",
                    (datetime.now().strftime('%Y-%m-%d %H:%M'), plant, 'Nowe', order_desc))
        order_id = cur.lastrowid
        for e in entries:
            cur.execute("""
            INSERT INTO clothes_order_items(order_id, worker_id, name, surname, item, size, qty, issued)
            VALUES(?,?,?,?,?,?,?,0)
            """, (order_id, e['worker_id'], e['name'], e['surname'], e['item'], e['size'], e['qty']))
        self.conn.commit()
        return order_id

    def _load_clothes_order_entries(self, order_id):
        rows = self.conn.execute("""
        SELECT COALESCE(coi.worker_id, 0),
               COALESCE(w.name, coi.name, ''),
               COALESCE(w.surname, coi.surname, ''),
               COALESCE(coi.item, ''),
               COALESCE(coi.size, ''),
               COALESCE(coi.qty, 1)
        FROM clothes_order_items coi
        LEFT JOIN workers w ON w.id = coi.worker_id
        WHERE coi.order_id=?
        ORDER BY COALESCE(w.surname, coi.surname, ''), COALESCE(w.name, coi.name, ''), COALESCE(coi.item, '')
        """, (order_id,)).fetchall()
        entries = []
        for r in rows:
            entries.append({
                'worker_id': r[0],
                'name': r[1],
                'surname': r[2],
                'item': r[3],
                'size': r[4],
                'qty': int(r[5] or 1),
            })
        return entries

    def generate_order_excels(self, order_id):
        entries = self._load_clothes_order_entries(order_id)
        if not entries:
            self.msg('Info', 'Brak pozycji do eksportu Excel')
            return
        p1, p2 = self._export_clothes_order_excels(order_id, entries)
        if p1 and p2:
            self.msg('OK', f"Excel wygenerowany.\n1) Hurtownia: {p1}\n2) Wydanie dla pracowników: {p2}")

    def _export_clothes_order_excels(self, order_id, entries):
        if Workbook is None:
            self.msg('Błąd', 'Brak openpyxl - nie można wygenerować raportów Excel')
            return None, None
        out_dir = Path('/storage/emulated/0/Documents/FutureExport') if platform == 'android' else Path('./exports')
        out_dir.mkdir(parents=True, exist_ok=True)

        summary = defaultdict(int)
        for e in entries:
            key = (e['item'], e.get('size') or '-')
            summary[key] += int(e.get('qty') or 0)

        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = 'Hurtownia'
        ws1.append(['Pozycja', 'Rozmiar', 'Ilość'])
        for (item, size), qty in sorted(summary.items(), key=lambda x: (x[0][0], x[0][1])):
            ws1.append([item, size, qty])
        try:
            self.style_xlsx(ws1)
        except Exception:
            pass
        p1 = out_dir / f'zamowienie_hurtownia_{order_id}.xlsx'
        wb1.save(p1)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = 'Wydanie'
        ws2.append(['Pracownik', 'Pozycja', 'Rozmiar', 'Ilość'])
        for e in sorted(entries, key=lambda x: (x['surname'], x['name'], x['item'])):
            ws2.append([f"{e['name']} {e['surname']}", e['item'], e.get('size') or '-', e['qty']])
        try:
            self.style_xlsx(ws2)
        except Exception:
            pass
        p2 = out_dir / f'raport_wydania_{order_id}.xlsx'
        wb2.save(p2)
        return str(p1), str(p2)

    def create_order_ui(self):
        workers = self._clothes_fetch_workers_for_order()
        if not workers:
            return self.msg('Info', 'Brak pracowników do zamówienia')

        root = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        root.add_widget(Label(text='Nowe zamówienie odzieży', bold=True, size_hint_y=None, height=dp(36)))

        def labeled_input(title, hint):
            wrap = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(82), spacing=dp(2))
            wrap.add_widget(Label(text=title, halign='left', size_hint_y=None, height=dp(20), color=(0.82,0.86,0.93,1)))
            ti = ModernInput(hint_text=hint, size_hint_y=None, height=dp(58))
            wrap.add_widget(ti)
            return wrap, ti

        plant_wrap, plant_ti = labeled_input('Zakład zamówienia', 'Zakład (filtr pracowników)')
        desc_wrap, desc_ti = labeled_input('Nazwa / opis zamówienia', 'Np. Zamówienie zimowe - brygada A')
        search_wrap, search_ti = labeled_input('Wyszukiwarka pracownika', 'Szukaj pracownika...')
        root.add_widget(plant_wrap)
        root.add_widget(desc_wrap)
        root.add_widget(search_wrap)

        workers_grid = GridLayout(cols=1, size_hint_y=None, spacing=dp(4))
        workers_grid.bind(minimum_height=workers_grid.setter('height'))
        rows_ui = []

        def add_worker_row(w):
            row = BoxLayout(size_hint_y=None, height=dp(42), spacing=dp(6))
            cb = CheckBox(size_hint_x=None, width=dp(40))
            txt = Label(text=f"{w['name']} {w['surname']} ({w['plant'] or '-'})", halign='left')
            txt.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            row.add_widget(txt)
            row.add_widget(cb)
            rows_ui.append((w, row, cb))
            workers_grid.add_widget(row)

        for w in workers:
            add_worker_row(w)

        sc = ScrollView()
        sc.add_widget(workers_grid)
        root.add_widget(sc)

        def refresh_filter(*_):
            workers_grid.clear_widgets()
            q = search_ti.text.lower().strip()
            plant_q = plant_ti.text.lower().strip()
            for w, row, cb in rows_ui:
                txt = f"{w['name']} {w['surname']} {w['plant']}".lower()
                if q and q not in txt:
                    continue
                if plant_q and plant_q not in (w['plant'] or '').lower() and plant_q not in txt:
                    continue
                workers_grid.add_widget(row)

        def select_all_visible(_):
            visible = set(workers_grid.children)
            for _w, row, cb in rows_ui:
                if row in visible:
                    cb.active = True

        def select_plant(_):
            ptxt = plant_ti.text.lower().strip()
            if not ptxt:
                return self.msg('Info', 'Podaj zakład, aby zaznaczyć wszystkich z zakładu')
            for w, _row, cb in rows_ui:
                cb.active = ptxt in (w['plant'] or '').lower()

        def next_step(_):
            chosen = [w for w, _row, cb in rows_ui if cb.active]
            if not chosen:
                return self.msg('Błąd', 'Wybierz co najmniej jednego pracownika')
            p.dismiss()
            self._create_order_items_ui(chosen, plant_ti.text.strip(), desc_ti.text.strip())

        search_ti.bind(text=refresh_filter)
        plant_ti.bind(text=refresh_filter)

        root.add_widget(ModernButton(text='Zapisz zamówienie', on_press=next_step, bg_color=(0.16,0.56,0.33,1), size_hint_y=None, height=dp(50), font_size='18sp'))

        p = Popup(title='Nowe zamówienie - wybór pracowników', content=root, size_hint=(0.95,0.95))
        p.open()

    def _create_order_items_ui(self, selected_workers, plant, order_desc):
        root = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
        root.add_widget(Label(text='Konfiguracja zamówienia (ilości per pracownik)', bold=True, size_hint_y=None, height=dp(34)))

        grid = GridLayout(cols=1, size_hint_y=None, spacing=dp(8))
        grid.bind(minimum_height=grid.setter('height'))
        worker_forms = {}

        for w in selected_workers:
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(320), padding=dp(8), spacing=dp(6))
            with card.canvas.before:
                Color(*COLOR_CARD)
                rr = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(10)])
            card.bind(pos=lambda inst, val, r=rr: setattr(r, 'pos', val))
            card.bind(size=lambda inst, val, r=rr: setattr(r, 'size', val))

            head = BoxLayout(size_hint_y=None, height=dp(36))
            cb = CheckBox(active=True, size_hint_x=None, width=dp(42))
            hl = Label(text=f"{w['name']} {w['surname']} ({w['plant'] or '-'})", halign='left')
            hl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            head.add_widget(hl)
            head.add_widget(cb)
            card.add_widget(head)

            qty_map = {}
            items_grid = GridLayout(cols=3, size_hint_y=None, row_default_height=dp(34), row_force_default=True)
            items_grid.height = dp(34 * 6)
            items_grid.add_widget(Label(text='Pozycja', bold=True))
            items_grid.add_widget(Label(text='Rozmiar', bold=True))
            items_grid.add_widget(Label(text='Ilość', bold=True))
            for item_name in ['Koszulka', 'Bluza', 'Spodnie', 'Kurtka', 'Buty']:
                size_txt = w['sizes'].get(item_name) or '-'
                qti = TextInput(text='1', multiline=False, input_filter='int')
                qty_map[item_name] = qti
                items_grid.add_widget(Label(text=item_name))
                items_grid.add_widget(Label(text=size_txt))
                items_grid.add_widget(qti)
            card.add_widget(items_grid)
            worker_forms[w['id']] = {'use': cb, 'qty': qty_map}
            grid.add_widget(card)

        sc = ScrollView()
        sc.add_widget(grid)
        root.add_widget(sc)

        def order_all(_):
            for frm in worker_forms.values():
                frm['use'].active = True
                for qti in frm['qty'].values():
                    if not qti.text.strip() or qti.text.strip() == '0':
                        qti.text = '1'

        def save_order(_):
            entries = self._collect_order_entries(selected_workers, worker_forms)
            if not entries:
                return self.msg('Błąd', 'Brak pozycji do zamówienia')
            order_id = self._save_clothes_order_entries(plant, order_desc, entries)
            p.dismiss()
            try:
                scr = self.clothes_sm.get_screen('orders')
                if hasattr(scr, 'refresh'):
                    scr.refresh()
            except Exception:
                pass
            self.msg('OK', f"Zamówienie #{order_id} zapisane")

        root.add_widget(ModernButton(text='Zapisz zamówienie', on_press=save_order, bg_color=(0.16,0.56,0.33,1), size_hint_y=None, height=dp(52), font_size='16sp'))

        p = Popup(title='Nowe zamówienie - pozycje i ilości', content=root, size_hint=(0.97,0.97))
        p.open()

    def clothes_order_details(self, order_id):
        cur = self.conn.cursor()
        order_meta = cur.execute("SELECT COALESCE(order_desc,''), COALESCE(status,''), COALESCE(plant,'') FROM clothes_orders WHERE id=?", (order_id,)).fetchone()
        order_desc = (order_meta[0] if order_meta else '') or ''
        order_status = (order_meta[1] if order_meta else '') or ''
        order_plant = (order_meta[2] if order_meta else '') or ''
        root = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
        root.add_widget(Label(text=f"Szczegóły zamówienia #{order_id}", bold=True, size_hint_y=None, height=dp(40)))
        if order_desc:
            root.add_widget(Label(text=f"Opis: {order_desc}", size_hint_y=None, height=dp(28), halign='left'))
        root.add_widget(Label(text=f"Zakład: {order_plant}    Status: {order_status}", size_hint_y=None, height=dp(28), halign='left'))
        grid = GridLayout(cols=1, size_hint_y=None, spacing=dp(6))
        grid.bind(minimum_height=grid.setter('height'))
        rows = cur.execute("""
        SELECT coi.id, coi.worker_id,
               COALESCE(w.name, coi.name, ''),
               COALESCE(w.surname, coi.surname, ''),
               coi.item,
               COALESCE(coi.size,''),
               coi.qty,
               COALESCE(coi.issued,0)
        FROM clothes_order_items coi
        LEFT JOIN workers w ON w.id=coi.worker_id
        WHERE coi.order_id=?
        """,(order_id,)).fetchall()
        for r in rows:
            cid, wid, name, surname, item, size, qty, issued = r
            row = BoxLayout(size_hint_y=None, height=dp(66), spacing=dp(8))
            worker = f"{name or ''} {surname or ''}".strip()
            lbl = Label(text=f"{worker} - {item} {size or '-'} x{qty} {'(wydane)' if issued else ''}", halign='left', valign='middle')
            lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(12), None)))
            row.add_widget(lbl)
            btns = BoxLayout(size_hint_x=None, width=dp(128), orientation='vertical', spacing=dp(6))
            btns.add_widget(ModernButton(text="Usuń", bg_color=(0.7,0.1,0.1,1), size_hint_y=None, height=dp(38), on_press=lambda x, cid=cid: self._remove_order_item_and_refresh(cid, order_id, p)))
            btns.add_widget(ModernButton(text="Wydaj", size_hint_y=None, height=dp(38), on_press=lambda x, cid=cid: self._issue_order_item_and_refresh(cid, order_id, p)))
            row.add_widget(btns)
            grid.add_widget(row)
        scroll = ScrollView()
        scroll.add_widget(grid)
        root.add_widget(scroll)
        bottom = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(64))
        bottom.add_action(ModernButton(text="Dodaj pozycję", on_press=lambda x: self._add_position_to_order_ui(order_id, p)))
        bottom.add_action(ModernButton(text="Generuj Excel", on_press=lambda x: self.generate_order_excels(order_id)))
        bottom.add_action(ModernButton(text="Zamów", on_press=lambda x: self.mark_order_ordered(order_id)))
        bottom.add_action(ModernButton(text="Wydaj częściowo", on_press=lambda x: self.clothes_issue_partial(order_id)))
        bottom.add_action(ModernButton(text="Wydaj wszystkie", on_press=lambda x: [self.clothes_issue_all(order_id), p.dismiss()]))
        root.add_widget(bottom)
        popup_title = f"Zamówienie #{order_id}"
        if order_desc:
            popup_title = f"{popup_title} - {order_desc}"
        p = Popup(title=popup_title, content=root, size_hint=(0.95,0.95))
        p.open()

    def _remove_order_item_and_refresh(self, cid, order_id, popup):
        try:
            self.conn.execute("DELETE FROM clothes_order_items WHERE id=?", (cid,))
            self.conn.commit()
            self.msg("OK", "Pozycja usunięta")
            popup.dismiss()
            try:
                scr = self.clothes_sm.get_screen('orders')
                if hasattr(scr, 'refresh'):
                    scr.refresh()
            except:
                pass
        except Exception as e:
            self.msg("Błąd", str(e))

    def _issue_order_item_and_refresh(self, cid, order_id, popup):
        try:
            if not self._can_issue_order(order_id):
                return
            cur = self.conn.cursor()
            cur.execute("""
            SELECT coi.worker_id,
                   COALESCE(w.name, coi.name, ''),
                   COALESCE(w.surname, coi.surname, ''),
                   coi.item,
                   COALESCE(coi.size, ''),
                   COALESCE(coi.issued,0)
            FROM clothes_order_items coi
            LEFT JOIN workers w ON w.id=coi.worker_id
            WHERE coi.id=?
            """, (cid,))
            r = cur.fetchone()
            if not r:
                self.msg("Błąd", "Brak pozycji")
                return
            wid, name, surname, item, size, issued = r
            if int(issued or 0) == 1:
                self.msg("Info", "Pozycja jest już wydana")
                return
            cur.execute("""
            INSERT INTO clothes_history(worker_id, name, surname, item, size, date)
            VALUES(?,?,?,?,?,?)
            """, (wid, name, surname, item, size, datetime.now().strftime("%Y-%m-%d")))
            cur.execute("UPDATE clothes_order_items SET issued=1 WHERE id=?", (cid,))
            self._refresh_order_issue_status(order_id)
            self.conn.commit()
            self.msg("OK", "Pozycja wydana")
            popup.dismiss()
            try:
                scr = self.clothes_sm.get_screen('orders')
                if hasattr(scr, 'refresh'):
                    scr.refresh()
            except:
                pass
        except Exception as e:
            self.msg("Błąd", str(e))

    def _add_position_to_order_ui(self, order_id, parent_popup=None):
        box = BoxLayout(orientation='vertical', padding=dp(10), spacing=dp(8))
        box.add_widget(Label(text="Dodaj pozycję do zamówienia", bold=True))
        workers_grid = GridLayout(cols=1, size_hint_y=None)
        workers_grid.bind(minimum_height=workers_grid.setter('height'))
        rows = self.conn.execute("SELECT id,name,surname,plant FROM workers ORDER BY surname").fetchall()
        sel = []
        for r in rows:
            cb = CheckBox(size_hint_x=None, width=dp(40))
            row = BoxLayout(size_hint_y=None, height=dp(36))
            row.add_widget(Label(text=f"{r[1]} {r[2]} ({r[3]})"))
            row.add_widget(cb)
            workers_grid.add_widget(row)
            sel.append((r[0], cb))
        scroll = ScrollView(size_hint=(1, None), size=(0, dp(140)))
        scroll.add_widget(workers_grid)
        box.add_widget(scroll)
        item_ti = ModernInput(hint_text="Nazwa pozycji")
        qty_ti = ModernInput(hint_text="Ilość", text="1")
        box.add_widget(item_ti)
        box.add_widget(qty_ti)
        def run(_):
            selected = [wid for wid,cb in sel if cb.active]
            if not selected:
                self.msg("Błąd", "Wybierz przynajmniej jednego pracownika")
                return
            itemname = item_ti.text.strip()
            try:
                qty = int(qty_ti.text.strip())
            except:
                qty = 1
            if not itemname:
                self.msg("Błąd", "Podaj nazwę pozycji")
                return
            cur = self.conn.cursor()
            for wid in selected:
                cur.execute("INSERT INTO clothes_order_items(order_id, worker_id, item, qty, issued) VALUES(?,?,?,?,?)",
                            (order_id, wid, itemname, qty, 0))
            self.conn.commit()
            self.msg("OK", "Pozycje dodane")
            add_popup.dismiss()
            if parent_popup:
                parent_popup.dismiss()
            try:
                scr = self.clothes_sm.get_screen('orders')
                if hasattr(scr, 'refresh'):
                    scr.refresh()
            except:
                pass
        box.add_widget(ModernButton(text="Dodaj", on_press=run))
        add_popup = Popup(title="Dodaj pozycję", content=box, size_hint=(0.9,0.9))
        add_popup.open()

    def mark_order_ordered(self, order_id):
        try:
            self.conn.execute("UPDATE clothes_orders SET status='Zamówione' WHERE id=? AND COALESCE(status,'') NOT IN ('Wydane','Częściowo wydane')", (order_id,))
            self.conn.commit()
            self.msg("OK", "Zmieniono status na 'Zamówione'")
            try:
                scr = self.clothes_sm.get_screen('orders')
                if hasattr(scr, 'refresh'):
                    scr.refresh()
            except:
                pass
        except Exception as e:
            self.msg("Błąd", str(e))

    def export_clothes_history_csv(self):
        try:
            p = Path(self.user_data_dir) / "clothes_history.csv"
            rows = self.conn.execute("SELECT worker_id, item, date FROM clothes_history ORDER BY date DESC").fetchall()
            with open(p, "w", encoding="utf-8") as f:
                f.write("worker_id,item,date\n")
                for r in rows:
                    f.write(f"{r[0]},{r[1]},{r[2]}\n")
            self.msg("OK", f"Zapisano {p.name}")
        except Exception as e:
            self.msg("Błąd", str(e))

    def form_clothes_size(self, record=None):
        box = BoxLayout(orientation="vertical", padding=dp(10), spacing=dp(8))

        def labeled_field(title, hint, text_val=""):
            wrap = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(82), spacing=dp(2))
            wrap.add_widget(Label(text=title, halign='left', size_hint_y=None, height=dp(20), color=(0.82,0.86,0.93,1)))
            ti = ModernInput(hint_text=hint, text=text_val, size_hint_y=None, height=dp(58))
            wrap.add_widget(ti)
            return wrap, ti

        fields = [
            ("Imię", "Imię", record[1] if record else ""),
            ("Nazwisko", "Nazwisko", record[2] if record else ""),
            ("Zakład", "Zakład", record[3] if record else ""),
            ("Rozmiar koszulki", "Koszulka", record[4] if record else ""),
            ("Rozmiar bluzy", "Bluza", record[5] if record else ""),
            ("Rozmiar spodni", "Spodnie", record[6] if record else ""),
            ("Rozmiar kurtki", "Kurtka", record[7] if record else ""),
            ("Rozmiar butów", "Buty", record[8] if record else ""),
        ]
        inputs = []
        for title, hint, txt in fields:
            wrap, ti = labeled_field(title, hint, txt)
            box.add_widget(wrap)
            inputs.append(ti)

        name_ti, surname_ti, plant_ti, shirt_ti, hoodie_ti, pants_ti, jacket_ti, shoes_ti = inputs

        def save(_):
            try:
                if record and record[0]:
                    self.conn.execute("""
                    UPDATE clothes_sizes SET name=?, surname=?, plant=?, shirt=?, hoodie=?, pants=?, jacket=?, shoes=? WHERE id=?
                    """, (name_ti.text.strip(), surname_ti.text.strip(), plant_ti.text.strip(), shirt_ti.text.strip(), hoodie_ti.text.strip(), pants_ti.text.strip(), jacket_ti.text.strip(), shoes_ti.text.strip(), record[0]))
                else:
                    self.conn.execute("""
                    INSERT INTO clothes_sizes (name,surname,plant,shirt,hoodie,pants,jacket,shoes) VALUES (?,?,?,?,?,?,?,?)
                    """, (name_ti.text.strip(), surname_ti.text.strip(), plant_ti.text.strip(), shirt_ti.text.strip(), hoodie_ti.text.strip(), pants_ti.text.strip(), jacket_ti.text.strip(), shoes_ti.text.strip()))
                self._sync_worker_to_contacts_and_sizes(
                    name_ti.text.strip(),
                    surname_ti.text.strip(),
                    "",
                    plant_ti.text.strip()
                )
                self.conn.commit()
                self.msg("OK", "Zapisano rozmiary")
                p.dismiss()
                try:
                    scr = self.clothes_sm.get_screen('sizes')
                    if hasattr(scr, 'refresh'):
                        scr.refresh()
                except:
                    pass
            except Exception as e:
                self.msg("Błąd", str(e))

        box.add_widget(ModernButton(text="ZAPISZ", on_press=save, size_hint_y=None, height=dp(52)))
        p = Popup(title="Rozmiary pracownika", content=box, size_hint=(0.92,0.95))
        p.open()

    def edit_clothes_size(self, record):
        self.form_clothes_size(record)

    def delete_clothes_size(self, rec_id):
        def do_delete(_):
            try:
                self.conn.execute("DELETE FROM clothes_sizes WHERE id=?", (rec_id,))
                self.conn.commit()
                self.msg("OK", "Usunięto rekord")
                px.dismiss()
                try:
                    scr = self.clothes_sm.get_screen('sizes')
                    if hasattr(scr, 'refresh'):
                        scr.refresh()
                except:
                    pass
            except Exception as e:
                self.msg("Błąd", str(e))
        px = Popup(title="Usuń?", content=BoxLayout(orientation="vertical", children=[ModernButton(text="USUŃ", on_press=do_delete, size_hint_y=None, height=dp(50))]), size_hint=(0.7,0.3))
        px.open()

    def process_excel(self, path):
        try:
            if str(path).endswith(".xls") and xlrd:
                wb = xlrd.open_workbook(path); ws = wb.sheet_by_index(0); raw = [[str(ws.cell_value(r,c)).strip() for c in range(ws.ncols)] for r in range(ws.nrows)]
            else:
                wb = load_workbook(path, data_only=True); ws = wb.active; raw = [["" if v is None else str(v).strip() for v in r] for r in ws.iter_rows(values_only=True)]
            h_idx = 0
            for i, r in enumerate(raw[:15]):
                if any(x in " ".join([str(v) for v in r]).lower() for x in ["imię", "imie", "nazwisko"]): h_idx = i; break
            self.full_data = raw[h_idx:]
            self.filtered_data = self.full_data
            self.export_indices = list(range(len(self.full_data[0])))
            for i,v in enumerate(self.full_data[0]):
                v = str(v).lower()
                if "imi" in v: self.idx_name = i
                if "naz" in v: self.idx_surname = i
                if "pesel" in v: self.idx_pesel = i
            self.msg("OK", "Arkusz wczytany")
            self.log(f"Loaded excel: {path}")
        except Exception as e:
            self.log(f"process_excel error: {traceback.format_exc()}")
            self.msg("BŁĄD", "Plik uszkodzony")

    def send_individual_from_table(self, row):
        name, sur = str(row[self.idx_name]).strip(), str(row[self.idx_surname]).strip()
        pes = str(row[self.idx_pesel]).strip() if self.idx_pesel != -1 else ""
        res = self.conn.execute("SELECT email FROM contacts WHERE pesel=? AND pesel != ''", (pes,)).fetchone() if pes else None
        if not res: res = self.conn.execute("SELECT email FROM contacts WHERE name=? AND surname=? COLLATE NOCASE", (name.lower(), sur.lower())).fetchone()
        if not res: return self.msg("Błąd", f"Brak maila dla: {name}")
        def task():
            cfg_p = Path(self.user_data_dir)/"smtp.json"
            if not cfg_p.exists(): return Clock.schedule_once(lambda d: self.msg("!", "Brak SMTP"))
            cfg = json.load(open(cfg_p)); srv = self.connect_smtp(cfg)
            if self.send_single_email(srv, cfg, row, res[0]): Clock.schedule_once(lambda d: self.msg("OK", f"Wysłano do: {name}"))
            srv.quit()
        threading.Thread(target=task, daemon=True).start()

    def setup_email_ui(self):
        self.sc_ref["email"].clear_widgets()
        shell = AppLayout(title="Moduł Email")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm, 'current', 'home')))
        shell.nav_tabs.add_action(SecondaryButton(text="SMTP", on_press=lambda x: setattr(self.sm, 'current', 'smtp')))

        body = BoxLayout(orientation="vertical", spacing=dp(10))
        auto_card = Card(orientation="horizontal", size_hint_y=None, height=dp(54), spacing=dp(10))
        self.cb_auto = CheckBox(size_hint_x=None, width=dp(45))
        self.cb_auto.active = self.auto_send_mode
        self.cb_auto.bind(active=self.on_auto_checkbox_changed)
        auto_card.add_widget(self.cb_auto)
        auto_card.add_widget(Label(text="AUTOMATYCZNA WYSYŁKA", bold=True))
        body.add_widget(auto_card)

        self.lbl_stats = Label(text="Baza: 0", size_hint_y=None, height=dp(34)); body.add_widget(self.lbl_stats)
        self.pb_label = Label(text="Gotowy", size_hint_y=None, height=dp(28)); self.pb = ProgressBar(max=100, size_hint_y=None, height=dp(24)); body.add_widget(self.pb_label); body.add_widget(self.pb)

        actions = AppActionBar()
        actions.add_action(DangerButton(text="Wyczyść załączniki", on_press=self.clear_all_attachments, size_hint_x=None))
        actions.add_action(PrimaryButton(text="Edytuj szablon", on_press=lambda x: setattr(self.sm, 'current', 'tmpl'), size_hint_x=None))
        actions.add_action(PrimaryButton(text="Dodaj załącznik", on_press=lambda x: self.open_picker("attachment"), size_hint_x=None))
        actions.add_action(PrimaryButton(text="Wyślij jeden plik", on_press=self.start_special_send_flow, size_hint_x=None))
        actions.add_action(PrimaryButton(text="Start masowa wysyłka", on_press=self.start_mass_mailing, size_hint_x=None))
        actions.add_action(SecondaryButton(text="Pauza/Resume", on_press=self.toggle_pause_mailing, size_hint_x=None))

        body.add_widget(actions)
        shell.set_content(body)
        self.sc_ref["email"].add_widget(shell)
        self.update_stats()

    def on_auto_checkbox_changed(self, instance, value):
        self.auto_send_mode = bool(value)
        try:
            if hasattr(self, 'cb_paski_auto') and self.cb_paski_auto.active != value:
                self.cb_paski_auto.active = value
        except: pass

    def _cell_str(self, row, idx):
        if idx == -1 or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    def _norm_header(self, txt):
        t = "" if txt is None else str(txt).strip().lower()
        repl = {
            'ą': 'a', 'ć': 'c', 'ę': 'e', 'ł': 'l', 'ń': 'n', 'ó': 'o', 'ś': 's', 'ż': 'z', 'ź': 'z'
        }
        for a, b in repl.items():
            t = t.replace(a, b)
        return " ".join(t.replace("_", " ").replace("-", " ").split())

    def _clean_excel_number_text(self, value):
        s = "" if value is None else str(value).strip()
        if s.endswith('.0'):
            core = s[:-2]
            if core.replace('-', '').isdigit():
                s = core
        return s

    def _normalize_phone(self, value):
        s = self._clean_excel_number_text(value)
        s = s.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        if not s:
            return ""
        if s.startswith('00'):
            s = '+' + s[2:]
        if s.startswith('+'):
            digits = ''.join(ch for ch in s[1:] if ch.isdigit())
            return f'+{digits}' if digits else ''
        digits = ''.join(ch for ch in s if ch.isdigit())
        return f'+{digits}' if digits else ''

    def _normalize_pesel(self, value):
        s = self._clean_excel_number_text(value)
        digits = ''.join(ch for ch in s if ch.isdigit())
        if not digits:
            return ""
        if len(digits) < 11:
            digits = digits.zfill(11)
        elif len(digits) > 11:
            digits = digits[-11:]
        return digits

    def _normalize_plant_name(self, value):
        s = self._clean_excel_number_text(value)
        return " ".join(s.split())


    def _find_header_row_and_map(self, rows):
        aliases = {
            'name': ['imie', 'imi', 'name', 'first name'],
            'surname': ['nazwisko', 'nazw', 'surname', 'last name'],
            'email': ['email', 'e mail', 'mail'],
            'pesel': ['pesel'],
            'phone': ['telefon', 'tel', 'phone', 'kom'],
            'plant': ['zaklad', 'zaklad pracy', 'plant', 'oddzial', 'dzial', 'pracownicy', 'firma', 'company', 'pracodawca'],
            'apartment': ['adres', 'mieszkanie', 'apart', 'lokal'],
            'notes': ['notat', 'uwag', 'opis', 'notes'],
            'shirt': ['koszul', 'shirt', 'tshirt', 't shirt'],
            'hoodie': ['bluza', 'hoodie', 'hood'],
            'pants': ['spodnie', 'spodn', 'pants', 'trous'],
            'jacket': ['kurtka', 'kurt', 'jacket'],
            'shoes': ['but', 'shoe', 'obuwie'],
            'car_name': ['samochod', 'nazwa auta', 'pojazd', 'model', 'car', 'auto'],
            'registration': ['rejestr', 'nr rej', 'tablica', 'plate', 'registration'],
            'driver': ['kierowca', 'driver'],
            'mileage': ['przebieg', 'km', 'mileage'],
            'service_interval': ['interwal', 'serwis co', 'service interval'],
            'last_service': ['ostatni serwis', 'last service'],
            'city': ['miasto', 'city'],
            'address': ['adres', 'address', 'ulica'],
            'plant_phone': ['telefon zakladu', 'kontakt', 'contact phone', 'phone zaklad']
        }

        max_scan = min(len(rows), 25)
        best_idx = -1
        best_score = 0
        best_map = {}

        for ridx in range(max_scan):
            row = rows[ridx]
            headers = [self._norm_header(v) for v in row]
            mapping = {}
            for field, keys in aliases.items():
                found = -1
                for cidx, hv in enumerate(headers):
                    if not hv:
                        continue
                    if any(k in hv for k in keys):
                        found = cidx
                        break
                mapping[field] = found

            score = sum(1 for v in mapping.values() if v != -1)
            if mapping['name'] != -1 and mapping['surname'] != -1:
                score += 6
            if mapping['registration'] != -1:
                score += 2
            if score > best_score:
                best_score = score
                best_idx = ridx
                best_map = mapping

        if best_score < 2:
            return -1, {}
        return best_idx, best_map

    def process_book(self, path):
        try:
            if load_workbook is None:
                self.msg("Błąd", "Brak openpyxl - import niemożliwy")
                return

            wb = load_workbook(path, data_only=True)
            if not wb.worksheets:
                self.msg("Błąd", "Pusty plik")
                return

            imported_contacts = 0
            imported_workers = 0
            imported_sizes = 0
            imported_cars = 0
            imported_plants = 0
            skipped_only_name_surname = 0
            touched_sheets = 0

            people = {}
            cars_map = {}
            plants_map = {}

            def merge_field(dst, key, val):
                if val and (not dst.get(key) or len(val) > len(dst.get(key, ''))):
                    dst[key] = val

            def to_int(v, default=0):
                try:
                    s = str(v).replace(' ', '').replace(',', '.').strip()
                    if not s:
                        return default
                    return int(float(s))
                except Exception:
                    return default

            for ws in wb.worksheets:
                raw = list(ws.iter_rows(values_only=True))
                if not raw:
                    continue

                h_idx, m = self._find_header_row_and_map(raw)
                if h_idx == -1:
                    continue

                touched_sheets += 1
                has_name_surname = m.get('name', -1) != -1 and m.get('surname', -1) != -1
                has_sizes = any(m.get(k, -1) != -1 for k in ('shirt', 'hoodie', 'pants', 'jacket', 'shoes'))
                has_car = m.get('registration', -1) != -1 or m.get('car_name', -1) != -1
                has_plant = m.get('plant', -1) != -1

                for r in raw[h_idx + 1:]:
                    try:
                        n = self._clean_excel_number_text(self._cell_str(r, m.get('name', -1)))
                        sname = self._clean_excel_number_text(self._cell_str(r, m.get('surname', -1)))
                        plant = self._normalize_plant_name(self._cell_str(r, m.get('plant', -1)))

                        if has_name_surname and n and sname:
                            email = self._clean_excel_number_text(self._cell_str(r, m.get('email', -1))).lower()
                            pesel = self._normalize_pesel(self._cell_str(r, m.get('pesel', -1)))
                            phone = self._normalize_phone(self._cell_str(r, m.get('phone', -1)))
                            apartment = self._clean_excel_number_text(self._cell_str(r, m.get('apartment', -1)))
                            notes = self._clean_excel_number_text(self._cell_str(r, m.get('notes', -1)))
                            shirt = self._clean_excel_number_text(self._cell_str(r, m.get('shirt', -1))) if has_sizes else ""
                            hoodie = self._clean_excel_number_text(self._cell_str(r, m.get('hoodie', -1))) if has_sizes else ""
                            pants = self._clean_excel_number_text(self._cell_str(r, m.get('pants', -1))) if has_sizes else ""
                            jacket = self._clean_excel_number_text(self._cell_str(r, m.get('jacket', -1))) if has_sizes else ""
                            shoes = self._clean_excel_number_text(self._cell_str(r, m.get('shoes', -1))) if has_sizes else ""

                            only_name_surname = not any([email, pesel, phone, plant, apartment, notes, shirt, hoodie, pants, jacket, shoes])
                            if only_name_surname:
                                skipped_only_name_surname += 1
                                continue

                            key = (n.strip().lower(), sname.strip().lower())
                            p = people.get(key)
                            if not p:
                                p = {
                                    'name': n.strip(), 'surname': sname.strip(), 'email': '', 'pesel': '', 'phone': '',
                                    'plant': '', 'apartment': '', 'notes': '',
                                    'shirt': '', 'hoodie': '', 'pants': '', 'jacket': '', 'shoes': ''
                                }
                                people[key] = p

                            merge_field(p, 'email', email)
                            merge_field(p, 'pesel', pesel)
                            merge_field(p, 'phone', phone)
                            merge_field(p, 'plant', plant)
                            merge_field(p, 'apartment', apartment)
                            merge_field(p, 'notes', notes)
                            merge_field(p, 'shirt', shirt)
                            merge_field(p, 'hoodie', hoodie)
                            merge_field(p, 'pants', pants)
                            merge_field(p, 'jacket', jacket)
                            merge_field(p, 'shoes', shoes)

                        if has_car:
                            reg = self._clean_excel_number_text(self._cell_str(r, m.get('registration', -1))).upper()
                            car_name = self._clean_excel_number_text(self._cell_str(r, m.get('car_name', -1)))
                            driver = self._clean_excel_number_text(self._cell_str(r, m.get('driver', -1)))
                            mileage_raw = self._clean_excel_number_text(self._cell_str(r, m.get('mileage', -1)))
                            interval_raw = self._clean_excel_number_text(self._cell_str(r, m.get('service_interval', -1)))
                            last_service_raw = self._clean_excel_number_text(self._cell_str(r, m.get('last_service', -1)))

                            if reg or car_name:
                                if not car_name:
                                    car_name = reg or "Auto"
                                if not reg:
                                    reg = car_name.upper()
                                cars_map[reg] = (
                                    car_name,
                                    driver,
                                    max(0, to_int(mileage_raw, 0)),
                                    max(1, to_int(interval_raw, 15000)),
                                    max(0, to_int(last_service_raw, 0))
                                )

                        if has_plant and plant:
                            city = self._clean_excel_number_text(self._cell_str(r, m.get('city', -1)))
                            address = self._clean_excel_number_text(self._cell_str(r, m.get('address', -1)))
                            plant_phone = self._normalize_phone(self._cell_str(r, m.get('plant_phone', -1)))
                            notes = self._clean_excel_number_text(self._cell_str(r, m.get('notes', -1)))
                            pkey = plant.lower()
                            cur = plants_map.get(pkey)
                            if not cur:
                                plants_map[pkey] = {
                                    'name': plant,
                                    'city': city,
                                    'address': address,
                                    'phone': plant_phone,
                                    'notes': notes
                                }
                            else:
                                if len(plant) > len(cur.get('name', '')):
                                    cur['name'] = plant
                                if city and not cur.get('city'):
                                    cur['city'] = city
                                if address and not cur.get('address'):
                                    cur['address'] = address
                                if plant_phone and not cur.get('phone'):
                                    cur['phone'] = plant_phone
                                if notes and not cur.get('notes'):
                                    cur['notes'] = notes
                    except Exception:
                        self.log(f"process_book row import error [{ws.title}]: {traceback.format_exc()}")

            if touched_sheets == 0:
                self.msg("Błąd", "Nie wykryto nagłówków danych w żadnym arkuszu")
                return

            existing_workers = {
                (str(n).strip().lower(), str(s).strip().lower()): wid
                for wid, n, s in self.conn.execute("SELECT id, name, surname FROM workers").fetchall()
            }
            existing_sizes = {
                (str(n).strip().lower(), str(s).strip().lower()): sid
                for sid, n, s in self.conn.execute("SELECT id, name, surname FROM clothes_sizes").fetchall()
            }
            existing_cars = {
                str(reg or '').strip().upper(): cid
                for cid, reg in self.conn.execute("SELECT id, registration FROM cars").fetchall()
                if str(reg or '').strip()
            }

            contacts_rows = []
            worker_updates = []
            worker_inserts = []
            size_updates = []
            size_inserts = []

            for p in people.values():
                key = (p['name'].lower(), p['surname'].lower())
                contacts_rows.append((p['name'].lower(), p['surname'].lower(), p['email'], p['pesel'], p['phone'], p['plant'], p['apartment'], p['notes']))

                wid = existing_workers.get(key)
                if wid:
                    worker_updates.append((p['plant'], p['phone'], wid))
                else:
                    worker_inserts.append((p['name'], p['surname'], p['plant'], p['phone'], "", ""))

                if any([p['shirt'], p['hoodie'], p['pants'], p['jacket'], p['shoes']]):
                    sid = existing_sizes.get(key)
                    if sid:
                        size_updates.append((p['plant'], p['shirt'], p['hoodie'], p['pants'], p['jacket'], p['shoes'], sid))
                    else:
                        size_inserts.append((p['name'], p['surname'], p['plant'], p['shirt'], p['hoodie'], p['pants'], p['jacket'], p['shoes']))

            if contacts_rows:
                self.conn.executemany(
                    "INSERT OR REPLACE INTO contacts (name,surname,email,pesel,phone,workplace,apartment,notes) VALUES (?,?,?,?,?,?,?,?)",
                    contacts_rows
                )
            if worker_updates:
                self.conn.executemany("UPDATE workers SET plant=?, phone=? WHERE id=?", worker_updates)
            if worker_inserts:
                self.conn.executemany(
                    "INSERT INTO workers(name, surname, plant, phone, position, hire_date) VALUES(?,?,?,?,?,?)",
                    worker_inserts
                )
            if size_updates:
                self.conn.executemany(
                    "UPDATE clothes_sizes SET plant=?, shirt=?, hoodie=?, pants=?, jacket=?, shoes=? WHERE id=?",
                    size_updates
                )
            if size_inserts:
                self.conn.executemany(
                    "INSERT INTO clothes_sizes (name, surname, plant, shirt, hoodie, pants, jacket, shoes) VALUES (?,?,?,?,?,?,?,?)",
                    size_inserts
                )

            car_updates = []
            car_inserts = []
            for reg, vals in cars_map.items():
                cid = existing_cars.get(reg)
                if cid:
                    car_updates.append((vals[0], vals[1], vals[2], vals[3], vals[4], cid))
                else:
                    car_inserts.append((vals[0], reg, vals[1], vals[2], vals[3], vals[4]))

            if car_updates:
                self.conn.executemany(
                    "UPDATE cars SET name=?, driver=?, mileage=?, service_interval=?, last_service=? WHERE id=?",
                    car_updates
                )
            if car_inserts:
                self.conn.executemany(
                    "INSERT INTO cars(name, registration, driver, mileage, service_interval, last_service) VALUES(?,?,?,?,?,?)",
                    car_inserts
                )

            if plants_map:
                self.conn.executemany(
                    "INSERT INTO plants(name, city, address, contact_phone, notes) VALUES(?,?,?,?,?) "
                    "ON CONFLICT(name) DO UPDATE SET city=excluded.city, address=excluded.address, contact_phone=excluded.contact_phone, notes=excluded.notes",
                    [(vals['name'], vals.get('city', ''), vals.get('address', ''), vals.get('phone', ''), vals.get('notes', '')) for vals in plants_map.values()]
                )

            imported_contacts = len(contacts_rows)
            imported_workers = len(worker_updates) + len(worker_inserts)
            imported_sizes = len(size_updates) + len(size_inserts)
            imported_cars = len(car_updates) + len(car_inserts)
            imported_plants = len(plants_map)

            self.conn.commit()
            self.sync_all_contact_links()
            self._increment_db_version()
            self.update_stats()
            self.msg(
                "OK",
                f"Import zakończony (arkusze: {touched_sheets}).\nKontakty: {imported_contacts}\nPracownicy: {imported_workers}\nRozmiary: {imported_sizes}\nAuta: {imported_cars}\nZakłady: {imported_plants}\nPominięto (tylko imię+nazwisko): {skipped_only_name_surname}"
            )
            self.log(
                f"Imported workbook(all sheets): {path} | sheets={touched_sheets} contacts={imported_contacts} workers={imported_workers} sizes={imported_sizes} cars={imported_cars} plants={imported_plants} skipped_only_name_surname={skipped_only_name_surname}"
            )
        except Exception as e:
            self.log(f"process_book error: {traceback.format_exc()}")
            self.msg("BŁĄD", f"Nieudany import: {str(e)[:120]}")

    def _sanitize_col(self, name):
        s = "".join(c if c.isalnum() else "_" for c in str(name).strip().lower())
        s = s.strip("_")
        if not s:
            s = "col"
        return s

    def _increment_db_version(self):
        cur = self.conn.execute("SELECT val FROM settings WHERE key='db_version'").fetchone()
        v = int(cur[0]) if cur and cur[0].isdigit() else 0
        v += 1
        self.conn.execute("INSERT OR REPLACE INTO settings VALUES (?,?)", ('db_version', str(v)))
        self.conn.commit()
        return v

    def mailing_worker(self):
        cfg_p = Path(self.user_data_dir)/"smtp.json"
        if not cfg_p.exists(): return self.finish_mailing("Brak SMTP")
        cfg = json.load(open(cfg_p)); b_on, b_sz, proc = cfg.get('batch', True), 30, 0
        try:
            srv = self.connect_smtp(cfg)
            while self.queue:
                if self.mailing_paused:
                    time.sleep(0.5)
                    continue
                row = self.queue.pop(0); n, s = str(row[self.idx_name]).strip(), str(row[self.idx_surname]).strip()
                p_exc = str(row[self.idx_pesel]).strip() if self.idx_pesel != -1 else ""
                res_p = self.conn.execute("SELECT email FROM contacts WHERE pesel=? AND pesel != ''", (p_exc,)).fetchone() if p_exc else None
                target, vrf = (res_p[0], False) if res_p else (None, False)
                if not target:
                    res_n = self.conn.execute("SELECT email FROM contacts WHERE name=? AND surname=? COLLATE NOCASE", (n.lower(), s.lower())).fetchone()
                    if res_n: target, vrf = res_n[0], not self.auto_send_mode
                if target:
                    if vrf:
                        self.wait_for_user = True; Clock.schedule_once(lambda dt: self.ask_before_send_worker(row, target, n, s))
                        while self.wait_for_user: time.sleep(0.5)
                        if self.user_decision == "skip": continue
                    if self.send_single_email(srv, cfg, row, target): self.stats["ok"] += 1; self.session_details.append(f"OK: {n} {s}")
                    else: self.stats["fail"] += 1; srv.quit(); srv = self.connect_smtp(cfg)
                    proc += 1
                    if self.queue:
                        if b_on and proc >= b_sz: srv.quit(); time.sleep(60); srv = self.connect_smtp(cfg); proc = 0
                        else: time.sleep(random.uniform(3, 7))
                else: self.stats["skip"] += 1; self.session_details.append(f"SKIP: {n} {s}")
                Clock.schedule_once(lambda dt: self.update_progress(self.total_q - len(self.queue)))
            srv.quit(); self.finish_mailing("Zakończono")
        except Exception as e:
            self.log(f"mailing_worker error: {traceback.format_exc()}")
            self.finish_mailing(f"Error: {e}")

    def connect_smtp(self, cfg):
        s = smtplib.SMTP(cfg.get('h','smtp.gmail.com'), int(cfg.get('port',587)), timeout=25); s.starttls(); s.login(cfg['u'], cfg['p']); return s

    def send_single_email(self, srv, cfg, row_data, target):
        try:
            nx, sx = str(row_data[self.idx_name]).title(), str(row_data[self.idx_surname]).title()
            msg = EmailMessage(); ts, tb = self.conn.execute("SELECT val FROM settings WHERE key='t_sub'").fetchone(), self.conn.execute("SELECT val FROM settings WHERE key='t_body'").fetchone()
            msg["Subject"] = (ts[0] if ts else "Raport").replace("{Imię}", nx); msg["From"], msg["To"] = cfg['u'], target
            msg.set_content((tb[0] if tb else "Dzień dobry").replace("{Imię}", nx).replace("{Data}", datetime.now().strftime("%d.%m.%Y")))
            t_f = Path(self.user_data_dir)/f"r_{nx}.xlsx"; wb = Workbook(); ws = wb.active
            ws.append([self.full_data[0][k] for k in self.export_indices]); ws.append([str(row_data[k]) if (str(row_data[k]).strip()!="") else "0" for k in self.export_indices])
            try:
                self.style_xlsx(ws)
            except:
                pass
            wb.save(t_f)
            with open(t_f, "rb") as f: msg.add_attachment(f.read(), maintype="application", subtype="xlsx", filename=f"Raport_{nx}_{sx}.xlsx")
            for p in self.global_attachments:
                if os.path.exists(p):
                    ct, _ = mimetypes.guess_type(p); mn, sb = (ct or 'application/octet-stream').split('/', 1)
                    with open(p,"rb") as f: msg.add_attachment(f.read(), maintype=mn, subtype=sb, filename=os.path.basename(p))
            srv.send_message(msg); return True
        except Exception:
            self.log(f"send_single_email error: {traceback.format_exc()}")
            return False

    def style_xlsx(self, ws):
        s, c = Side(style='thin'), Alignment(horizontal='center', vertical='center')
        for ri, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                cell.border = Border(top=s, left=s, right=s, bottom=s); cell.alignment = c
                if ri == 1: cell.font = Font(bold=True); cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                elif ri % 2 == 0: cell.fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
        for col in ws.columns:
            m = 0; col_let = col[0].column_letter
            for cell in col:
                if cell.value: m = max(m, len(str(cell.value)))
            ws.column_dimensions[col_let].width = (m * 1.3) + 7

    def setup_smtp_ui(self):
        self.sc_ref["smtp"].clear_widgets()
        p = Path(self.user_data_dir)/"smtp.json"
        d = json.load(open(p)) if p.exists() else {}

        shell = AppLayout(title="Ustawienia SMTP")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm,'current','home')))

        form = BoxLayout(orientation="vertical", spacing=dp(10))
        self.ti_h = ModernInput(hint_text="Host", text=d.get('h',''))
        self.ti_pt = ModernInput(hint_text="Port", text=str(d.get('port','587')))
        self.ti_u = ModernInput(hint_text="Email/Login", text=d.get('u',''))
        self.ti_p = ModernInput(hint_text="Hasło/Klucz", password=True, text=d.get('p',''))
        form.add_widget(self.ti_h); form.add_widget(self.ti_pt); form.add_widget(self.ti_u); form.add_widget(self.ti_p)

        bx = Card(orientation="horizontal", size_hint_y=None, height=dp(52), spacing=dp(10))
        self.cb_b = CheckBox(size_hint_x=None, width=dp(45), active=d.get('batch', True))
        bx.add_widget(self.cb_b); bx.add_widget(Label(text="Batching (przerwa 60s/30 maili)"))
        form.add_widget(bx)

        actions = AppActionBar()
        actions.add_action(PrimaryButton(text="Zapisz", on_press=lambda x: [json.dump({'h':self.ti_h.text,'port':self.ti_pt.text,'u':self.ti_u.text,'p':self.ti_p.text,'batch':self.cb_b.active}, open(p,"w")), self.msg("OK","Zapisano")], size_hint_x=None))
        actions.add_action(PrimaryButton(text="Test połączenia", on_press=lambda x: self.test_smtp_direct(), size_hint_x=None))
        actions.add_action(SecondaryButton(text="Pokaż logi", on_press=self.show_logs, size_hint_x=None))

        body = BoxLayout(orientation="vertical", spacing=dp(10))
        body.add_widget(form)
        body.add_widget(actions)
        shell.set_content(body)
        self.sc_ref["smtp"].add_widget(shell)

    def test_smtp_direct(self):
        try: s = self.connect_smtp({'h':self.ti_h.text,'port':self.ti_pt.text,'u':self.ti_u.text,'p':self.ti_p.text}); s.quit(); self.msg("OK", "Serwer SMTP Działa!"); self.log("SMTP test succeeded")
        except Exception as e: self.log(f"test_smtp_direct error: {traceback.format_exc()}"); self.msg("BŁĄD", str(e)[:60])

    def start_special_send_flow(self, _): self.open_picker("special_send")

    def special_send_step_2(self, path):
        self.selected_emails = []; box = BoxLayout(orientation="vertical", padding=dp(15), spacing=dp(10)); ti = ModernInput(hint_text="Szukaj..."); box.add_widget(ti)
        sc = ScrollView(); gl = GridLayout(cols=1, size_hint_y=None, spacing=dp(5)); gl.bind(minimum_height=gl.setter('height')); sc.add_widget(gl); box.add_widget(sc)
        def rf(v=""):
            gl.clear_widgets(); rows = self.conn.execute("SELECT name, surname, email FROM contacts").fetchall()
            for r in rows:
                if v and v.lower() not in f"{r[0]} {r[1]} {r[2]}".lower(): continue
                bx = BoxLayout(size_hint_y=None, height=dp(50)); cb = CheckBox(size_hint_x=None, width=dp(50))
                cb.bind(active=lambda inst, val, m=r[2]: self.selected_emails.append(m) if val else self.selected_emails.remove(m))
                bx.add_widget(cb); bx.add_widget(Label(text=f"{r[0].title()} {r[1].title()}")); gl.add_widget(bx)
        ti.bind(text=lambda i,v: rf(v)); rf(); btn = ModernButton(text="DALEJ", on_press=lambda x: [p.dismiss(), self.special_send_step_3(path)] if self.selected_emails else None); box.add_widget(btn); p = Popup(title="Odbiorcy", content=box, size_hint=(.95,.9)); p.open()

    def special_send_step_3(self, path):
        b = BoxLayout(orientation="vertical", padding=dp(15), spacing=dp(10)); ti_s = ModernInput(hint_text="Temat"); ti_b = ModernInput(hint_text="Treść", multiline=True); b.add_widget(ti_s); b.add_widget(ti_b)
        def run(_):
            def task():
                cfg = json.load(open(Path(self.user_data_dir)/"smtp.json")); srv = self.connect_smtp(cfg)
                for m in self.selected_emails:
                    msg = EmailMessage(); msg["Subject"], msg["From"], msg["To"] = ti_s.text, cfg['u'], m; msg.set_content(ti_b.text)
                    with open(path, "rb") as f: msg.add_attachment(f.read(), maintype="application", subtype="octet-stream", filename=os.path.basename(path))
                    srv.send_message(msg)
                srv.quit(); Clock.schedule_once(lambda d: self.msg("OK", "Wysłano")); self.log(f"Special send file {path} to {len(self.selected_emails)} recipients")
            threading.Thread(target=task, daemon=True).start(); p.dismiss()
        b.add_widget(ModernButton(text="WYŚLIJ PLIK", on_press=run)); p = Popup(title="Wiadomość", content=b, size_hint=(.9, .8)); p.open()

    def filter_table(self, i, v): self.filtered_data = [self.full_data[0]] + [r for r in self.full_data[1:] if any(v.lower() in str(c).lower() for c in r)]; self.refresh_table()

    def start_mass_mailing(self, _):
        if self.is_mailing_running: return
        self.stats, self.session_details, self.queue = {"ok": 0, "fail": 0, "skip": 0}, [], list(self.full_data[1:])
        self.total_q = len(self.queue); self.is_mailing_running = True; self.mailing_paused = False; threading.Thread(target=self.mailing_worker, daemon=True).start()
        self.log(f"Mass mailing started: {self.total_q} items")

    def open_picker(self, mode):
        if platform != "android": return self.msg("!", "Tylko Android")
        from jnius import autoclass; from android import activity
        PA, Intent = autoclass("org.kivy.android.PythonActivity"), autoclass("android.content.Intent"); intent = Intent(Intent.ACTION_GET_CONTENT); intent.setType("*/*")
        if mode == "attachment": intent.putExtra(Intent.EXTRA_ALLOW_MULTIPLE, True)
        def cb(req, res, dt):
            if req != 1001: return
            activity.unbind(on_activity_result=cb)
            if res == -1 and dt:
                resolver = PA.mActivity.getContentResolver(); files = []
                clip = dt.getClipData()
                if clip: [files.append(clip.getItemAt(i).getUri()) for i in range(clip.getItemCount())]
                else: files.append(dt.getData())
                for uri in files:
                    cur = resolver.query(uri, None, None, None, None); name = f"f_{random.randint(10,99)}.xlsx"
                    if cur and cur.moveToFirst(): idx = cur.getColumnIndex("_display_name"); name = cur.getString(idx) if idx != -1 else name; cur.close()
                    try:
                        stream, loc = resolver.openInputStream(uri), Path(self.user_data_dir) / name
                        with open(loc, "wb") as f:
                            buf = bytearray(16384)
                            while (n := stream.read(buf)) > 0: f.write(buf[:n])
                        stream.close()
                        if mode == "data": self.process_excel(loc)
                        elif mode == "book": self.process_book(loc)
                        elif mode == "attachment":
                            self.global_attachments.append(str(loc))
                        elif mode == "special_send": Clock.schedule_once(lambda dt: self.special_send_step_2(str(loc)))
                    except:
                        pass
                self.update_stats()
        activity.bind(on_activity_result=cb); PA.mActivity.startActivityForResult(intent, 1001)

    def setup_tmpl_ui(self):
        self.sc_ref["tmpl"].clear_widgets()
        ts = self.conn.execute("SELECT val FROM settings WHERE key='t_sub'").fetchone()
        tb = self.conn.execute("SELECT val FROM settings WHERE key='t_body'").fetchone()

        shell = AppLayout(title="Szablon email")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm, 'current', 'email')))

        form = BoxLayout(orientation="vertical", spacing=dp(10))
        ti_s = ModernInput(hint_text="Temat {Imię}")
        ti_b = ModernInput(hint_text="Treść...", multiline=True)
        ti_s.text, ti_b.text = (ts[0] if ts else ""), (tb[0] if tb else "")
        form.add_widget(ti_s)
        form.add_widget(ti_b)

        actions = AppActionBar()
        actions.add_action(PrimaryButton(text="Zapisz", on_press=lambda x: [self.conn.execute("INSERT OR REPLACE INTO settings VALUES (?,?)", ('t_sub',ti_s.text)), self.conn.execute("INSERT OR REPLACE INTO settings VALUES (?,?)", ('t_body',ti_b.text)), self.conn.commit(), self.msg("OK","Wzór zapisany")], size_hint_x=None))

        body = BoxLayout(orientation="vertical", spacing=dp(10))
        body.add_widget(form)
        body.add_widget(actions)
        shell.set_content(body)
        self.sc_ref["tmpl"].add_widget(shell)

    def setup_contacts_ui(self):
        self.sc_ref["contacts"].clear_widgets()
        shell = AppLayout(title="Kontakty")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm, 'current', 'home')))
        shell.nav_tabs.add_action(PrimaryButton(text="Dodaj", on_press=lambda x: self.form_contact(), size_hint_x=None, width=dp(150)))

        body = BoxLayout(orientation="vertical", spacing=dp(8))
        search_row = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(8))
        self.ti_cs = ModernInput(hint_text="Szukaj po imieniu, nazwisku, email, telefonie...")
        self.ti_cs.bind(text=self.refresh_contacts_list)
        search_row.add_widget(self.ti_cs)

        filter_row = BoxLayout(size_hint_y=None, height=dp(54), spacing=dp(8))
        self.ti_cs_workplace = ModernInput(hint_text="Filtr zakład pracy")
        self.ti_cs_workplace.bind(text=self.refresh_contacts_list)
        self.ti_cs_city = ModernInput(hint_text="Filtr adres / mieszkanie")
        self.ti_cs_city.bind(text=self.refresh_contacts_list)
        filter_row.add_widget(self.ti_cs_workplace)
        filter_row.add_widget(self.ti_cs_city)

        self.c_ls = GridLayout(cols=1, size_hint_y=None, spacing=dp(10), padding=[dp(2), dp(2)])
        self.c_ls.bind(minimum_height=self.c_ls.setter('height'))
        sc = ScrollView()
        sc.add_widget(self.c_ls)

        body.add_widget(search_row)
        body.add_widget(filter_row)
        body.add_widget(sc)
        shell.set_content(body)
        shell.set_fab(lambda x: self.form_contact())
        self.sc_ref["contacts"].add_widget(shell)

    def refresh_contacts_list(self, *args):
        self.c_ls.clear_widgets(); sv = self.ti_cs.text.lower()
        sv_workplace = self.ti_cs_workplace.text.lower() if hasattr(self, 'ti_cs_workplace') else ""
        sv_city = self.ti_cs_city.text.lower() if hasattr(self, 'ti_cs_city') else ""
        rows = self.conn.execute("SELECT name, surname, email, pesel, phone, workplace, apartment, notes FROM contacts ORDER BY surname ASC").fetchall()
        for d in rows:
            searchable = f"{d[0]} {d[1]} {d[2]} {d[4]} {d[5]} {d[6]} {d[7]}".lower()
            if sv and sv not in searchable:
                continue
            if sv_workplace and sv_workplace not in str(d[5]).lower():
                continue
            if sv_city and sv_city not in str(d[6]).lower():
                continue

            card = BoxLayout(orientation="vertical", size_hint_y=None, height=dp(250), padding=dp(10), spacing=dp(8))
            with card.canvas.before:
                Color(*COLOR_CARD)
                rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(12)])
            card.bind(pos=lambda inst, val, r=rect: setattr(r, 'pos', val), size=lambda inst, val, r=rect: setattr(r, 'size', val))

            name_lbl = Label(text=f"{d[0]} {d[1]}".title(), bold=True, halign="left", valign='middle', size_hint_y=None, height=dp(38))
            name_lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(6), None)))
            card.add_widget(name_lbl)

            info_text = (
                f"E: {d[2]}\n"
                f"PESEL: {d[3] if d[3] else '-'}\n"
                f"T: {d[4] if d[4] else '-'}\n"
                f"Zakład: {d[5] if d[5] else '-'}\n"
                f"Adres: {d[6] if d[6] else '-'}\n"
                f"Notatka: {d[7] if d[7] else '-'}"
            )
            info_lbl = Label(text=info_text, font_size='12sp', halign="left", valign='top', color=(0.84,0.86,0.92,1))
            info_lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(6), None)))
            card.add_widget(info_lbl)

            actions = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(132), min_button_height=dp(44))
            phone_txt = str(d[4]).strip() if d[4] else ""
            actions.add_action(ModernButton(text="Zadzwoń", on_press=lambda x, ph=phone_txt: self._call_contact(ph), bg_color=(0.16,0.6,0.3,1)))
            actions.add_action(ModernButton(text="WhatsApp", on_press=lambda x, ph=phone_txt, nm=d[0]: self._whatsapp_contact(ph, nm), bg_color=(0.06,0.55,0.25,1)))
            actions.add_action(ModernButton(text="Edytuj", on_press=lambda x, data=d: self.form_contact(*data)))
            actions.add_action(ModernButton(text="Usuń", bg_color=(0.8,0.2,0.2,1), on_press=lambda x, n=d[0], sn=d[1]: self.delete_contact(n, sn)))
            card.add_widget(actions)
            self.c_ls.add_widget(card)

    def msg(self, tit, txt):
        b = BoxLayout(orientation="vertical", padding=dp(18), spacing=dp(10))
        l = Label(text=txt, halign="center", valign="middle")
        l.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(8), None)))
        b.add_widget(l)
        b.add_widget(PrimaryButton(text="OK", on_press=lambda x: p.dismiss(), height=dp(54), size_hint_y=None))
        p = Popup(title=tit, content=b, size_hint=(0.92, 0.55), auto_dismiss=False)
        p.open()

    def update_stats(self, *a):
        try:
            count = self.conn.execute('SELECT count(*) FROM contacts').fetchone()[0]
            s = f"Baza: {count} | Załączniki: {len(self.global_attachments)}"
            if hasattr(self, 'lbl_stats'):
                self.lbl_stats.text = s
            if hasattr(self, 'lbl_stats_paski'):
                self.lbl_stats_paski.text = s
        except:
            pass

    def update_progress(self, d):
        try:
            val = int((d/self.total_q)*100) if self.total_q else 0
            if hasattr(self, 'pb'):
                self.pb.value = val
            if hasattr(self, 'pb_paski'):
                self.pb_paski.value = val
            if hasattr(self, 'pb_label'):
                self.pb_label.text = f"Postęp: {d}/{self.total_q}"
            if hasattr(self, 'pb_label_paski'):
                self.pb_label_paski.text = f"Postęp: {d}/{self.total_q}"
        except:
            pass

    def finish_mailing(self, s):
        self.is_mailing_running = False; det = "\n".join(self.session_details); self.conn.execute("INSERT INTO reports (date, ok, fail, skip, auto, details) VALUES (?,?,?,?,?,?)", (datetime.now().strftime("%Y-%m-%d %H:%M"), self.stats['ok'], self.stats['fail'], self.stats['skip'], 0, det)); self.conn.commit()
        Clock.schedule_once(lambda dt: self.msg("Mailing", f"{s}\nSukces: {self.stats['ok']}"))
        self.log(f"Mailing finished: {s} | ok={self.stats['ok']} fail={self.stats['fail']} skip={self.stats['skip']}")

    def popup_columns(self, _):
        box, gr, checks = BoxLayout(orientation="vertical", padding=dp(10)), GridLayout(cols=1, size_hint_y=None, spacing=dp(5)), []
        gr.bind(minimum_height=gr.setter('height'))
        for i, h in enumerate(self.full_data[0]):
            r, cb = BoxLayout(size_hint_y=None, height=dp(45)), CheckBox(active=(i in self.export_indices), size_hint_x=None, width=dp(50)); checks.append((i, cb)); r.add_widget(cb); r.add_widget(Label(text=str(h))); gr.add_widget(r)
        sc = ScrollView(); sc.add_widget(gr); box.add_widget(sc); box.add_widget(ModernButton(text="ZASTOSUJ", on_press=lambda x: [setattr(self, 'export_indices', [idx for idx, c in checks if c.active]), p.dismiss(), self.refresh_table()], height=dp(50), size_hint_y=None)); p = Popup(title="Kolumny", content=box, size_hint=(0.9, 0.9)); p.open()

    def setup_report_ui(self):
        self.sc_ref["report"].clear_widgets()
        shell = AppLayout(title="Historia sesji")
        shell.nav_tabs.add_action(SecondaryButton(text="Wróć", on_press=lambda x: setattr(self.sm, 'current', 'home')))

        self.r_grid = GridLayout(cols=1, size_hint_y=None, spacing=dp(10), padding=[dp(2), dp(2)])
        self.r_grid.bind(minimum_height=self.r_grid.setter('height'))
        sc = ScrollView(); sc.add_widget(self.r_grid)

        shell.set_content(sc)
        self.sc_ref["report"].add_widget(shell)

    def refresh_reports(self, *a):
        self.r_grid.clear_widgets(); rows = self.conn.execute("SELECT date, ok, fail, skip, details FROM reports ORDER BY id DESC").fetchall()
        for d, ok, fl, sk, det in rows:
            row = Card(orientation="vertical", size_hint_y=None, height=dp(120), padding=dp(10), spacing=dp(8))
            row.add_widget(Label(text=f"Sesja: {d}", bold=True, color=COLOR_PRIMARY))
            row.add_widget(Label(text=f"OK: {ok}  BŁĘDY: {fl}  POMINIĘTE: {sk}", color=(0.8,0.85,0.92,1), size_hint_y=None, height=dp(26)))
            row.add_widget(PrimaryButton(text="Pokaż logi", size_hint_y=None, height=dp(42), on_press=lambda x, t=det: self.show_details(t)))
            self.r_grid.add_widget(row)

    def show_details(self, t):
        b = BoxLayout(orientation="vertical", padding=dp(10)); ti = TextInput(text=str(t), readonly=True, font_size='11sp'); b.add_widget(ti); b.add_widget(Button(text="ZAMKNIJ", size_hint_y=0.2, on_press=lambda x: p.dismiss())); p = Popup(title="Logi", content=b, size_hint=(.9,.8)); p.open()

    def ask_before_send_worker(self, row, email, n, s):
        def dec(v): self.user_decision = "send" if v else "skip"; self.wait_for_user = False; px.dismiss()
        box = BoxLayout(orientation="vertical", padding=dp(20), spacing=dp(10)); box.add_widget(Label(text=f"POTWIERDŹ:\n[b]{n} {s}[/b]\n{email}", markup=True, halign="center"))
        btns = BoxLayout(size_hint_y=None, height=dp(55), spacing=dp(10)); btns.add_widget(Button(text="WYŚLIJ", on_press=lambda x: dec(True), background_color=(0,0.6,0,1))); btns.add_widget(Button(text="POMIŃ", on_press=lambda x: dec(False), background_color=(0.7,0,0,1)))
        box.add_widget(btns); px = Popup(title="Weryfikacja", content=box, size_hint=(0.9, 0.45), auto_dismiss=False); px.open()

    def export_single_row(self, r):
        p = Path("/storage/emulated/0/Documents/FutureExport") if platform=="android" else Path("./exports"); p.mkdir(parents=True, exist_ok=True)
        nx, sx = str(r[self.idx_name]).title(), str(r[self.idx_surname]).title(); wb = Workbook(); ws = wb.active
        ws.append([self.full_data[0][k] for k in self.export_indices]); ws.append([str(r[k]) if (k < len(r) and str(r[k]).strip() != "") else "0" for k in self.export_indices])
        try:
            self.style_xlsx(ws)
        except:
            pass
        wb.save(p/f"Raport_{nx}_{sx}.xlsx"); self.msg("OK", f"Zapisano PDF dla: {nx}"); self.log(f"Export single row for {nx} {sx}")

    def sync_all_contact_links(self):
        rows = self.conn.execute("SELECT name, surname, phone, workplace FROM contacts").fetchall()
        for n, s, ph, wp in rows:
            self._sync_contact_to_workers_and_sizes(n, s, ph, wp)
        self.conn.commit()

    def _sync_contact_to_workers_and_sizes(self, name, surname, phone="", workplace=""):
        n = str(name).strip()
        s = str(surname).strip()
        if not n or not s:
            return
        self.ensure_extended_tables()
        row = self.conn.execute(
            "SELECT id FROM workers WHERE lower(name)=lower(?) AND lower(surname)=lower(?) LIMIT 1",
            (n, s)
        ).fetchone()
        if row:
            self.conn.execute(
                "UPDATE workers SET plant=?, phone=? WHERE id=?",
                (str(workplace).strip(), str(phone).strip(), row[0])
            )
        else:
            self.conn.execute(
                "INSERT INTO workers(name, surname, plant, phone, position, hire_date) VALUES(?,?,?,?,?,?)",
                (n, s, str(workplace).strip(), str(phone).strip(), "", "")
            )

        sz = self.conn.execute(
            "SELECT id FROM clothes_sizes WHERE lower(name)=lower(?) AND lower(surname)=lower(?) LIMIT 1",
            (n, s)
        ).fetchone()
        if sz:
            self.conn.execute(
                "UPDATE clothes_sizes SET plant=COALESCE(NULLIF(?, ''), plant) WHERE id=?",
                (str(workplace).strip(), sz[0])
            )
        else:
            self.conn.execute(
                "INSERT INTO clothes_sizes(name, surname, plant, shirt, hoodie, pants, jacket, shoes) VALUES(?,?,?,?,?,?,?,?)",
                (n, s, str(workplace).strip(), "", "", "", "", "")
            )

    def _sync_worker_to_contacts_and_sizes(self, name, surname, phone="", plant=""):
        n = str(name).strip()
        s = str(surname).strip()
        if not n or not s:
            return
        existing = self.conn.execute(
            "SELECT email, pesel, apartment, notes FROM contacts WHERE lower(name)=lower(?) AND lower(surname)=lower(?) LIMIT 1",
            (n, s)
        ).fetchone()
        if existing:
            self.conn.execute(
                "UPDATE contacts SET phone=?, workplace=? WHERE lower(name)=lower(?) AND lower(surname)=lower(?)",
                (str(phone).strip(), str(plant).strip(), n, s)
            )
        else:
            self.conn.execute(
                "INSERT INTO contacts(name, surname, email, pesel, phone, workplace, apartment, notes) VALUES(?,?,?,?,?,?,?,?)",
                (n.lower(), s.lower(), "", "", str(phone).strip(), str(plant).strip(), "", "")
            )
        self._sync_contact_to_workers_and_sizes(n, s, phone, plant)

    def _bind_rect(self, widget, rect):
        widget.bind(pos=lambda inst, val, r=rect: setattr(r, 'pos', val))
        widget.bind(size=lambda inst, val, r=rect: setattr(r, 'size', val))

    def delete_contact(self, n, s):
        def pr(_):
            self.conn.execute("DELETE FROM contacts WHERE name=? AND surname=?", (n, s))
            self.conn.execute("DELETE FROM workers WHERE lower(name)=lower(?) AND lower(surname)=lower(?)", (n, s))
            self.conn.execute("DELETE FROM clothes_sizes WHERE lower(name)=lower(?) AND lower(surname)=lower(?)", (n, s))
            self.conn.commit()
            px.dismiss()
            self.refresh_contacts_list()
            self.refresh_workers_module()
            self.update_stats()
        px = Popup(title="Usuń?", content=BoxLayout(orientation="vertical", children=[ModernButton(text="USUŃ KONTAKT", on_press=pr, size_hint_y=None, height=dp(50))]), size_hint=(0.7,0.3)); px.open()

    def form_contact(self, n="", s="", e="", pes="", ph="", workplace="", apartment="", notes=""):
        b, f_ins = BoxLayout(orientation="vertical", padding=dp(15), spacing=dp(10)), [TextInput(text=str(n), hint_text="Imię"), TextInput(text=str(s), hint_text="Nazwisko"), TextInput(text=str(e), hint_text="Email"), TextInput(text=str(pes), hint_text="PESEL"), TextInput(text=str(ph), hint_text="Telefon")]
        for f in f_ins: b.add_widget(f)
        workplace_ti = TextInput(hint_text="Zakład pracy (np. Rybnik KWK Jankowice)", size_hint_y=None, height=dp(40), text=str(workplace))
        apartment_ti = TextInput(hint_text="Mieszkanie / adres", size_hint_y=None, height=dp(40), text=str(apartment))
        notes_ti = TextInput(hint_text="Notatki o kontakcie", size_hint_y=None, height=dp(70), multiline=True, text=str(notes))
        b.add_widget(workplace_ti); b.add_widget(apartment_ti); b.add_widget(notes_ti)
        def save(_):
            if not f_ins[0].text.strip() or not f_ins[1].text.strip():
                return self.msg("Błąd", "Imię i nazwisko są wymagane")
            self.conn.execute("INSERT OR REPLACE INTO contacts (name,surname,email,pesel,phone,workplace,apartment,notes) VALUES (?,?,?,?,?,?,?,?)",
                (f_ins[0].text.lower(),
                 f_ins[1].text.lower(),
                 f_ins[2].text.strip(),
                 f_ins[3].text.strip(),
                 f_ins[4].text.strip(),
                 workplace_ti.text.strip(),
                 apartment_ti.text.strip(),
                 notes_ti.text.strip()))
            self._sync_contact_to_workers_and_sizes(
                f_ins[0].text.strip(),
                f_ins[1].text.strip(),
                f_ins[4].text.strip(),
                workplace_ti.text.strip()
            )
            self.conn.commit()
            px.dismiss()
            self.refresh_contacts_list()
            self.refresh_workers_module()
            self.update_stats()
        b.add_widget(ModernButton(text="ZAPISZ", on_press=save)); px = Popup(title="Kontakt", content=b, size_hint=(0.9, 0.85)); px.open()

    def _normalize_phone(self, phone):
        raw = ''.join(ch for ch in str(phone or '') if ch.isdigit() or ch == '+')
        if raw.startswith('00'):
            return '+' + raw[2:]
        if raw and not raw.startswith('+') and len(raw) >= 9:
            return '+48' + raw[-9:]
        return raw

    def _call_contact(self, phone):
        ph = self._normalize_phone(phone)
        if not ph:
            return self.msg("Info", "Brak numeru telefonu")
        try:
            if platform == "android":
                from jnius import autoclass
                PA = autoclass("org.kivy.android.PythonActivity")
                Intent = autoclass("android.content.Intent")
                Uri = autoclass("android.net.Uri")
                intent = Intent(Intent.ACTION_DIAL)
                intent.setData(Uri.parse(f"tel:{ph}"))
                PA.mActivity.startActivity(intent)
            else:
                webbrowser.open(f"tel:{ph}")
        except Exception:
            self.msg("Błąd", "Nie udało się uruchomić dialera")

    def _whatsapp_contact(self, phone, name=""):
        ph = self._normalize_phone(phone).replace('+', '')
        if not ph:
            return self.msg("Info", "Brak numeru telefonu")
        text = urllib.parse.quote(f"Dzień dobry {str(name).title()}, ")
        url = f"https://wa.me/{ph}?text={text}"
        try:
            webbrowser.open(url)
        except Exception:
            self.msg("Błąd", "Nie udało się otworzyć WhatsApp")


    def clear_all_attachments(self, _):
        [self.global_attachments.clear(), self.update_stats(), self.log("Cleared attachments")]

    def refresh_cars_list(self, *args):
        """Odświeża listę samochodów na ekranie cars."""
        if not hasattr(self, 'cars_grid'):
            return
        self.cars_grid.clear_widgets()

        search = self.ti_cars_search.text.lower().strip() if hasattr(self, 'ti_cars_search') else ""
        self.init_cars_db()
        rows = self.conn.execute(
            "SELECT id, name, registration, driver, mileage, service_interval, last_service FROM cars ORDER BY name, registration"
        ).fetchall()

        for row in rows:
            car_id, name, registration, driver, mileage, service_interval, last_service = row
            blob = f"{name} {registration} {driver}".lower()
            if search and search not in blob:
                continue

            mileage = int(mileage or 0)
            service_interval = int(service_interval or 0)
            last_service = int(last_service or 0)
            remaining = service_interval - (mileage - last_service)
            remaining_txt = f"Do serwisu: {remaining} km"
            remaining_color = (0.9, 0.25, 0.25, 1) if remaining < 1500 else (0.78, 0.81, 0.87, 1)

            # karta samochodu: czytelna pionowa struktura + akcje w scrollu
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(222), padding=dp(10), spacing=dp(8))
            with card.canvas.before:
                Color(*COLOR_CARD)
                rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(12)])
            card.bind(pos=lambda inst, val, r=rect: setattr(r, 'pos', val), size=lambda inst, val, r=rect: setattr(r, 'size', val))

            title = Label(text=f"{name or '-'} | {registration or '-'}", bold=True, halign='left', size_hint_y=None, height=dp(36))
            title.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            card.add_widget(title)

            d1 = Label(text=f"Kierowca: {driver or '-'}", halign='left', size_hint_y=None, height=dp(26))
            d1.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            d2 = Label(text=f"Przebieg: {mileage} km | Interwał: {service_interval} km", halign='left', size_hint_y=None, height=dp(26))
            d2.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            d3 = Label(text=remaining_txt, halign='left', color=remaining_color, size_hint_y=None, height=dp(26))
            d3.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            card.add_widget(d1); card.add_widget(d2); card.add_widget(d3)

            actions = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(62), min_button_width=dp(170), min_button_height=dp(44))
            actions.add_action(ModernButton(text='Zmień kierowcę', on_press=lambda x, cid=car_id, cur=driver: self.change_driver_popup(cid, cur)))
            actions.add_action(ModernButton(text='Dodaj przebieg', on_press=lambda x, cid=car_id, cur=mileage: self.add_mileage_popup(cid, cur)))
            actions.add_action(ModernButton(text='Potwierdź serwis', on_press=lambda x, cid=car_id: self.confirm_service(cid), bg_color=(0.18,0.58,0.36,1)))
            actions.add_action(ModernButton(text='Usuń samochód', on_press=lambda x, cid=car_id: self.delete_car(cid), bg_color=(0.74,0.14,0.14,1)))
            card.add_widget(actions)

            self.cars_grid.add_widget(card)

    def _open_driver_picker(self, target_input):
        """Wybór kierowcy z wyszukiwarką opartą o kontakty i pracowników."""
        box = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        search = ModernInput(hint_text='Szukaj kierowcy (kontakty/pracownicy)...')
        box.add_widget(search)

        sc = ScrollView()
        gl = GridLayout(cols=1, size_hint_y=None, spacing=dp(6), padding=[dp(2), dp(2)])
        gl.bind(minimum_height=gl.setter('height'))
        sc.add_widget(gl)
        box.add_widget(sc)

        # Zbieramy kandydatów z obu modułów.
        people = set()
        try:
            for n, sn in self.conn.execute("SELECT name, surname FROM contacts").fetchall():
                full = f"{str(n or '').strip().title()} {str(sn or '').strip().title()}".strip()
                if full:
                    people.add(full)
        except Exception:
            pass
        try:
            for n, sn in self.conn.execute("SELECT name, surname FROM workers").fetchall():
                full = f"{str(n or '').strip().title()} {str(sn or '').strip().title()}".strip()
                if full:
                    people.add(full)
        except Exception:
            pass

        people = sorted(people)

        def refill(query=''):
            gl.clear_widgets()
            q = (query or '').lower().strip()
            shown = 0
            for person in people:
                if q and q not in person.lower():
                    continue
                shown += 1
                gl.add_widget(ModernButton(
                    text=person,
                    size_hint_y=None,
                    height=dp(46),
                    on_press=lambda x, p=person: [setattr(target_input, 'text', p), px.dismiss()]
                ))
            if shown == 0:
                gl.add_widget(Label(text='Brak wyników', size_hint_y=None, height=dp(34), color=(0.8,0.82,0.88,1)))

        search.bind(text=lambda inst, val: refill(val))
        refill()

        btns = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(58), min_button_width=dp(120))
        btns.add_action(ModernButton(text='Zamknij', on_press=lambda x: px.dismiss(), bg_color=(0.35,0.35,0.42,1)))
        box.add_widget(btns)

        px = Popup(title='Wybierz kierowcę', content=box, size_hint=(0.92, 0.8), auto_dismiss=False)
        px.open()

    def add_car_popup(self):
        """Popup dodawania nowego samochodu."""
        box = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        ti_name = ModernInput(hint_text='Nazwa samochodu')
        ti_reg = ModernInput(hint_text='Rejestracja')
        ti_driver = ModernInput(hint_text='Kierowca (opcjonalnie)')
        ti_int = ModernInput(hint_text='Interwał serwisowy (km)', text='15000')
        box.add_widget(ti_name); box.add_widget(ti_reg); box.add_widget(ti_driver); box.add_widget(ti_int)
        box.add_widget(ModernButton(text='Wybierz kierowcę z kontaktów/pracowników', size_hint_y=None, height=dp(44), on_press=lambda x: self._open_driver_picker(ti_driver)))

        def save(_):
            name = ti_name.text.strip()
            reg = ti_reg.text.strip().upper()
            try:
                interval = int(ti_int.text.strip() or '0')
            except Exception:
                return self.msg('Błąd', 'Interwał musi być liczbą')
            if not name or not reg:
                return self.msg('Błąd', 'Nazwa i rejestracja są wymagane')

            self.init_cars_db()
            self.conn.execute(
                "INSERT INTO cars(name, registration, driver, mileage, service_interval, last_service) VALUES(?,?,?,?,?,?)",
                (name, reg, ti_driver.text.strip(), 0, max(1, interval), 0)
            )
            self.conn.commit()
            px.dismiss()
            self.refresh_cars_list()

        btns = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(140))
        btns.add_action(ModernButton(text='Zapisz', on_press=save))
        btns.add_action(ModernButton(text='Anuluj', on_press=lambda x: px.dismiss(), bg_color=(0.35,0.35,0.42,1)))
        box.add_widget(btns)

        px = Popup(title='+ DODAJ SAMOCHÓD', content=box, size_hint=(0.92, 0.52), auto_dismiss=False)
        px.open()

    def change_driver_popup(self, car_id, current_driver=''):
        """Popup zmiany kierowcy."""
        box = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        ti_driver = ModernInput(hint_text='Imię kierowcy', text=str(current_driver or ''))
        box.add_widget(ti_driver)
        box.add_widget(ModernButton(text='Wybierz z kontaktów/pracowników', size_hint_y=None, height=dp(44), on_press=lambda x: self._open_driver_picker(ti_driver)))

        def save(_):
            self.conn.execute('UPDATE cars SET driver=? WHERE id=?', (ti_driver.text.strip(), car_id))
            self.conn.commit()
            px.dismiss()
            self.refresh_cars_list()

        btns = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(140))
        btns.add_action(ModernButton(text='Zapisz', on_press=save))
        btns.add_action(ModernButton(text='Anuluj', on_press=lambda x: px.dismiss(), bg_color=(0.35,0.35,0.42,1)))
        box.add_widget(btns)

        px = Popup(title='Zmień kierowcę', content=box, size_hint=(0.9, 0.4), auto_dismiss=False)
        px.open()

    def add_mileage_popup(self, car_id, current_mileage=0):
        """Popup aktualizacji przebiegu."""
        box = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        ti_mileage = ModernInput(hint_text='Nowy przebieg', text=str(current_mileage or 0))
        box.add_widget(ti_mileage)

        def save(_):
            try:
                new_m = int(ti_mileage.text.strip() or '0')
            except Exception:
                return self.msg('Błąd', 'Przebieg musi być liczbą')
            self.conn.execute('UPDATE cars SET mileage=? WHERE id=?', (max(0, new_m), car_id))
            self.conn.commit()
            px.dismiss()
            self.refresh_cars_list()

        btns = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(140))
        btns.add_action(ModernButton(text='Zapisz', on_press=save))
        btns.add_action(ModernButton(text='Anuluj', on_press=lambda x: px.dismiss(), bg_color=(0.35,0.35,0.42,1)))
        box.add_widget(btns)

        px = Popup(title='Dodaj przebieg', content=box, size_hint=(0.9, 0.4), auto_dismiss=False)
        px.open()

    def confirm_service(self, car_id):
        """Potwierdza serwis: last_service = mileage."""
        self.conn.execute('UPDATE cars SET last_service=mileage WHERE id=?', (car_id,))
        self.conn.commit()
        self.refresh_cars_list()
        self.msg('OK', 'Serwis został potwierdzony')

    def form_car(self, cid=None, plate='', brand='', model='', plant='', mileage=0, status='Aktywny', driver='', notes=''):
        """Kompatybilność: przekierowanie do nowego popupu dodawania samochodu."""
        self.add_car_popup()

    def delete_car(self, cid):
        """Usuwa samochód z tabeli cars."""
        self.conn.execute('DELETE FROM cars WHERE id=?', (cid,))
        self.conn.commit()
        self.refresh_cars_list()

    def refresh_workers_module(self, *args):
        if not hasattr(self, 'workers_grid'):
            return
        self.workers_grid.clear_widgets()
        search = self.ti_workers_search.text.lower() if hasattr(self, 'ti_workers_search') else ''
        rows = self.conn.execute('SELECT id, name, surname, plant, phone, position, hire_date FROM workers ORDER BY surname, name').fetchall()
        for row in rows:
            if search and search not in " ".join(str(x or '') for x in row).lower():
                continue
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(220), padding=dp(10), spacing=dp(8))
            with card.canvas.before:
                Color(*COLOR_CARD)
                rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(12)])
            card.bind(pos=lambda inst, val, r=rect: setattr(r, 'pos', val), size=lambda inst, val, r=rect: setattr(r, 'size', val))

            name_lbl = Label(text=f"{row[1] or '-'} {row[2] or '-'}", bold=True, halign='left', size_hint_y=None, height=dp(38))
            name_lbl.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            card.add_widget(name_lbl)
            d1 = Label(text=f"Stanowisko: {row[5] or '-'} | Zakład: {row[3] or '-'}", font_size='12sp', halign='left', size_hint_y=None, height=dp(28))
            d1.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            d2 = Label(text=f"Telefon: {row[4] or '-'} | Zatrudniony: {row[6] or '-'}", font_size='12sp', halign='left', color=(0.78,0.81,0.87,1), size_hint_y=None, height=dp(28))
            d2.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            card.add_widget(d1); card.add_widget(d2)

            actions = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(132), min_button_height=dp(44))
            actions.add_action(ModernButton(text='Edytuj', on_press=lambda x, data=row: self.form_worker(*data)))
            actions.add_action(ModernButton(text='Usuń', bg_color=(0.7,0.15,0.15,1), on_press=lambda x, wid=row[0]: self.delete_worker(wid)))
            card.add_widget(actions)
            self.workers_grid.add_widget(card)

    def form_worker(self, wid=None, name='', surname='', plant='', phone='', position='', hire_date=''):
        b = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        fields = {
            'name': TextInput(text=str(name or ''), hint_text='Imię'),
            'surname': TextInput(text=str(surname or ''), hint_text='Nazwisko'),
            'plant': TextInput(text=str(plant or ''), hint_text='Zakład'),
            'phone': TextInput(text=str(phone or ''), hint_text='Telefon'),
            'position': TextInput(text=str(position or ''), hint_text='Stanowisko'),
            'hire_date': TextInput(text=str(hire_date or ''), hint_text='Data zatrudnienia (YYYY-MM-DD)'),
        }
        for key in ['name', 'surname', 'plant', 'phone', 'position', 'hire_date']:
            b.add_widget(fields[key])

        def save(_):
            if not fields['name'].text.strip() or not fields['surname'].text.strip():
                return self.msg('Błąd', 'Imię i nazwisko są wymagane')
            if wid:
                self.conn.execute('UPDATE workers SET name=?, surname=?, plant=?, phone=?, position=?, hire_date=? WHERE id=?',
                    (fields['name'].text.strip(), fields['surname'].text.strip(), fields['plant'].text.strip(), fields['phone'].text.strip(), fields['position'].text.strip(), fields['hire_date'].text.strip(), wid))
            else:
                self.conn.execute('INSERT INTO workers(name, surname, plant, phone, position, hire_date) VALUES(?,?,?,?,?,?)',
                    (fields['name'].text.strip(), fields['surname'].text.strip(), fields['plant'].text.strip(), fields['phone'].text.strip(), fields['position'].text.strip(), fields['hire_date'].text.strip()))
            self._sync_worker_to_contacts_and_sizes(
                fields['name'].text.strip(),
                fields['surname'].text.strip(),
                fields['phone'].text.strip(),
                fields['plant'].text.strip()
            )
            self.conn.commit()
            px.dismiss()
            self.refresh_workers_module()

        b.add_widget(ModernButton(text='Zapisz', on_press=save))
        px = Popup(title='Pracownik', content=b, size_hint=(0.9, 0.85))
        px.open()

    def delete_worker(self, wid):
        self.conn.execute('DELETE FROM workers WHERE id=?', (wid,))
        self.conn.commit()
        self.refresh_workers_module()

    def refresh_plants_list(self, *args):
        if not hasattr(self, 'plants_grid'):
            return
        self.plants_grid.clear_widgets()
        search = self.ti_plants_search.text.lower() if hasattr(self, 'ti_plants_search') else ''
        try:
            rows = self.conn.execute('SELECT id, name, city, address, contact_phone, notes FROM plants ORDER BY name').fetchall()
        except Exception:
            self.ensure_extended_tables()
            rows = self.conn.execute('SELECT id, name, city, address, contact_phone, notes FROM plants ORDER BY name').fetchall()
        for row in rows:
            if search and search not in " ".join(str(x or '') for x in row).lower():
                continue
            card = BoxLayout(orientation='vertical', size_hint_y=None, height=dp(220), padding=dp(10), spacing=dp(8))
            with card.canvas.before:
                Color(*COLOR_CARD)
                rect = RoundedRectangle(pos=card.pos, size=card.size, radius=[dp(12)])
            card.bind(pos=lambda inst, val, r=rect: setattr(r, 'pos', val), size=lambda inst, val, r=rect: setattr(r, 'size', val))

            h = Label(text=f"{row[1] or '-'} ({row[2] or '-'})", bold=True, halign='left', size_hint_y=None, height=dp(36))
            h.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            i1 = Label(text=f"Adres: {row[3] or '-'}", font_size='12sp', halign='left', size_hint_y=None, height=dp(28))
            i1.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            i2 = Label(text=f"Tel: {row[4] or '-'} | Notatki: {row[5] or '-'}", font_size='12sp', halign='left', color=(0.78,0.81,0.87,1), size_hint_y=None, height=dp(28))
            i2.bind(size=lambda inst, val: setattr(inst, 'text_size', (inst.width - dp(4), None)))
            card.add_widget(h); card.add_widget(i1); card.add_widget(i2)

            actions = ButtonContainer(orientation='horizontal', size_hint_y=None, height=dp(60), min_button_width=dp(132), min_button_height=dp(44))
            actions.add_action(ModernButton(text='Edytuj', on_press=lambda x, data=row: self.form_plant(*data)))
            actions.add_action(ModernButton(text='Usuń', bg_color=(0.7,0.15,0.15,1), on_press=lambda x, pid=row[0]: self.delete_plant(pid)))
            card.add_widget(actions)
            self.plants_grid.add_widget(card)

    def form_plant(self, pid=None, name='', city='', address='', contact_phone='', notes=''):
        b = BoxLayout(orientation='vertical', padding=dp(12), spacing=dp(8))
        fields = {
            'name': TextInput(text=str(name or ''), hint_text='Nazwa zakładu'),
            'city': TextInput(text=str(city or ''), hint_text='Miasto'),
            'address': TextInput(text=str(address or ''), hint_text='Adres'),
            'contact_phone': TextInput(text=str(contact_phone or ''), hint_text='Telefon'),
            'notes': TextInput(text=str(notes or ''), hint_text='Notatki', multiline=True, size_hint_y=None, height=dp(70)),
        }
        for key in ['name', 'city', 'address', 'contact_phone', 'notes']:
            b.add_widget(fields[key])

        def save(_):
            if not fields['name'].text.strip():
                return self.msg('Błąd', 'Nazwa zakładu jest wymagana')
            if pid:
                self.conn.execute('UPDATE plants SET name=?, city=?, address=?, contact_phone=?, notes=? WHERE id=?',
                    (fields['name'].text.strip(), fields['city'].text.strip(), fields['address'].text.strip(), fields['contact_phone'].text.strip(), fields['notes'].text.strip(), pid))
            else:
                self.conn.execute('INSERT INTO plants(name, city, address, contact_phone, notes) VALUES(?,?,?,?,?)',
                    (fields['name'].text.strip(), fields['city'].text.strip(), fields['address'].text.strip(), fields['contact_phone'].text.strip(), fields['notes'].text.strip()))
            self.conn.commit()
            px.dismiss()
            self.refresh_plants_list()

        b.add_widget(ModernButton(text='Zapisz', on_press=save))
        px = Popup(title='Zakład', content=b, size_hint=(0.9, 0.85))
        px.open()

    def delete_plant(self, pid):
        self.conn.execute('DELETE FROM plants WHERE id=?', (pid,))
        self.conn.commit()
        self.refresh_plants_list()

    def setup_cars_ui(self):
        """Buduje ekran cars: nagłówek, lista i panel akcji."""
        self.sc_ref["cars"].clear_widgets()
        self.init_cars_db()

        shell = AppLayout(title="Samochody")
        shell.nav_tabs.add_action(SecondaryButton(text='Powrót', on_press=lambda x: setattr(self.sm, 'current', 'home')))
        shell.nav_tabs.add_action(PrimaryButton(text='+ DODAJ SAMOCHÓD', on_press=lambda x: self.add_car_popup(), size_hint_x=None, width=dp(210)))

        body = BoxLayout(orientation='vertical', spacing=dp(8))
        self.ti_cars_search = ModernInput(hint_text='Szukaj: nazwa / rejestracja / kierowca')
        self.ti_cars_search.bind(text=self.refresh_cars_list)
        body.add_widget(self.ti_cars_search)

        self.cars_grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None, padding=[dp(2), dp(2)])
        self.cars_grid.bind(minimum_height=self.cars_grid.setter('height'))
        sc = ScrollView()
        sc.add_widget(self.cars_grid)
        body.add_widget(sc)

        shell.set_content(body)
        shell.set_fab(lambda x: self.add_car_popup())
        self.sc_ref['cars'].add_widget(shell)
        self.refresh_cars_list()

    def setup_pracownicy_ui(self):
        self.sc_ref["pracownicy"].clear_widgets()
        shell = AppLayout(title="Pracownicy")
        shell.nav_tabs.add_action(SecondaryButton(text='Powrót', on_press=lambda x: setattr(self.sm, 'current', 'home')))
        shell.nav_tabs.add_action(PrimaryButton(text='Dodaj', on_press=lambda x: self.form_worker(), size_hint_x=None, width=dp(150)))
        body = BoxLayout(orientation='vertical', spacing=dp(8))
        self.ti_workers_search = ModernInput(hint_text='Szukaj pracownika (imię, nazwisko, zakład)')
        self.ti_workers_search.bind(text=self.refresh_workers_module)
        body.add_widget(self.ti_workers_search)
        self.workers_grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        self.workers_grid.bind(minimum_height=self.workers_grid.setter('height'))
        sc = ScrollView(); sc.add_widget(self.workers_grid)
        body.add_widget(sc)
        shell.set_content(body)
        shell.set_fab(lambda x: self.form_worker())
        self.sc_ref['pracownicy'].add_widget(shell)
        self.refresh_workers_module()

    def setup_zaklady_ui(self):
        self.sc_ref["zaklady"].clear_widgets()
        shell = AppLayout(title="Zakłady")
        shell.nav_tabs.add_action(SecondaryButton(text='Powrót', on_press=lambda x: setattr(self.sm, 'current', 'home')))
        shell.nav_tabs.add_action(PrimaryButton(text='Dodaj', on_press=lambda x: self.form_plant(), size_hint_x=None, width=dp(150)))
        body = BoxLayout(orientation='vertical', spacing=dp(8))
        self.ti_plants_search = ModernInput(hint_text='Szukaj zakładu (nazwa, miasto, telefon)')
        self.ti_plants_search.bind(text=self.refresh_plants_list)
        body.add_widget(self.ti_plants_search)
        self.plants_grid = GridLayout(cols=1, spacing=dp(8), size_hint_y=None)
        self.plants_grid.bind(minimum_height=self.plants_grid.setter('height'))
        sc = ScrollView(); sc.add_widget(self.plants_grid)
        body.add_widget(sc)
        shell.set_content(body)
        shell.set_fab(lambda x: self.form_plant())
        self.sc_ref['zaklady'].add_widget(shell)
        self.refresh_plants_list()

    def setup_settings_ui(self):
        self.sc_ref["settings"].clear_widgets()
        shell = AppLayout(title="Ustawienia i narzędzia")
        shell.nav_tabs.add_action(SecondaryButton(text="Powrót", on_press=lambda x: setattr(self.sm, 'current', 'home')))

        body = BoxLayout(orientation="vertical", spacing=dp(10))
        try:
            contacts_count = self.conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
            workers_count = self.conn.execute("SELECT COUNT(*) FROM workers").fetchone()[0]
            cars_count = self.conn.execute("SELECT COUNT(*) FROM fleet_cars").fetchone()[0]
            plants_count = self.conn.execute("SELECT COUNT(*) FROM plants").fetchone()[0]
            body.add_widget(Label(text=f"Baza: kontakty {contacts_count} | pracownicy {workers_count} | auta {cars_count} | zakłady {plants_count}", size_hint_y=None, height=dp(34), color=(0.75,0.82,0.92,1)))
        except Exception:
            pass

        actions = ScrollView()
        action_grid = GridLayout(cols=1, spacing=dp(10), size_hint_y=None, padding=[dp(2), dp(2)])
        action_grid.bind(minimum_height=action_grid.setter('height'))
        action_grid.add_widget(PrimaryButton(text="Dodaj bazę danych", on_press=lambda x: self.open_picker("book"), height=dp(54), size_hint_y=None))
        action_grid.add_widget(PrimaryButton(text="Ustawienia SMTP", on_press=lambda x: setattr(self.sm, 'current', 'smtp'), height=dp(54), size_hint_y=None))
        action_grid.add_widget(PrimaryButton(text="Edytuj szablon email", on_press=lambda x: setattr(self.sm, 'current', 'tmpl'), height=dp(54), size_hint_y=None))
        action_grid.add_widget(PrimaryButton(text="Wczytaj arkusz płac", on_press=lambda x: self.open_picker("data"), height=dp(54), size_hint_y=None))
        action_grid.add_widget(SecondaryButton(text="Pokaż logi", on_press=self.show_logs, height=dp(54), size_hint_y=None))
        actions.add_widget(action_grid)
        body.add_widget(actions)
        shell.set_content(body)
        self.sc_ref["settings"].add_widget(shell)

    def setup_paski_ui(self):
        self.sc_ref["paski"].clear_widgets()
        shell = AppLayout(title="Moduł Paski")
        shell.nav_tabs.add_action(SecondaryButton(text="Powrót", on_press=lambda x: setattr(self.sm, 'current', 'home')))

        body = BoxLayout(orientation="vertical", spacing=dp(10))
        auto_row = Card(orientation="horizontal", size_hint_y=None, height=dp(52), spacing=dp(10))
        self.cb_paski_auto = CheckBox(size_hint_x=None, width=dp(45))
        self.cb_paski_auto.active = self.auto_send_mode
        self.cb_paski_auto.bind(active=self.on_auto_checkbox_changed)
        auto_row.add_widget(self.cb_paski_auto)
        auto_row.add_widget(Label(text="AUTOMATYCZNA WYSYŁKA", bold=True))
        body.add_widget(auto_row)

        self.lbl_stats_paski = Label(text="Baza: 0 | Załączniki: 0", size_hint_y=None, height=dp(32)); body.add_widget(self.lbl_stats_paski)
        self.pb_label_paski = Label(text="Gotowy", size_hint_y=None, height=dp(28)); self.pb_paski = ProgressBar(max=100, size_hint_y=None, height=dp(24)); body.add_widget(self.pb_label_paski); body.add_widget(self.pb_paski)

        actions = AppActionBar()
        actions.add_action(PrimaryButton(text="Wczytaj arkusz płac", on_press=lambda x: self.open_picker("data"), size_hint_x=None))
        actions.add_action(PrimaryButton(text="Podgląd i eksport", on_press=lambda x: [self.refresh_table(), setattr(self.sm, 'current', 'table')] if self.full_data else self.msg("!", "Wczytaj arkusz!"), size_hint_x=None))
        actions.add_action(PrimaryButton(text="Edytuj szablon", on_press=lambda x: setattr(self.sm, 'current', 'tmpl'), size_hint_x=None))
        actions.add_action(PrimaryButton(text="Dołącz załącznik", on_press=lambda x: self.open_picker("attachment"), size_hint_x=None))
        actions.add_action(PrimaryButton(text="Wyślij jeden plik", on_press=self.start_special_send_flow, size_hint_x=None))
        actions.add_action(PrimaryButton(text="Start masowa wysyłka", on_press=self.start_mass_mailing, size_hint_x=None))
        actions.add_action(SecondaryButton(text="PAUZA/RESUME", on_press=self.toggle_pause_mailing, size_hint_x=None))
        actions.add_action(SecondaryButton(text="Raporty sesji", on_press=lambda x: [self.refresh_reports(), setattr(self.sm, 'current', 'report')], size_hint_x=None))
        actions.add_action(DangerButton(text="Wyczyść załączniki", on_press=self.clear_all_attachments, size_hint_x=None))

        body.add_widget(actions)
        shell.set_content(body)
        self.sc_ref["paski"].add_widget(shell)
        self.update_stats()

    def toggle_pause_mailing(self, _=None):
        self.mailing_paused = not self.mailing_paused
        if self.mailing_paused: self.msg("OK", "Wysyłka zatrzymana"); self.log("Mailing paused")
        else: self.msg("OK", "Wysyłka wznowiona"); self.log("Mailing resumed")

    def show_logs(self, _=None):
        try:
            text = ""
            if self.log_file.exists():
                with open(self.log_file, "r", encoding="utf-8") as f:
                    text = f.read()[-40000:]
            else:
                text = "\n".join(self._log_buffer)
            b = BoxLayout(orientation="vertical", padding=dp(10))
            ti = TextInput(text=text, readonly=True, font_size='11sp')
            b.add_widget(ti)
            b.add_widget(Button(text="ZAMKNIJ", size_hint_y=0.2, on_press=lambda x: p.dismiss()))
            p = Popup(title="Logi aplikacji", content=b, size_hint=(.95,.95)); p.open()
        except Exception:
            self.log(f"show_logs error: {traceback.format_exc()}")
            self.msg("Błąd", "Nie można otworzyć logów")

if __name__ == "__main__":
    FutureApp().run()
